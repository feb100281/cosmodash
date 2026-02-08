import pandas as pd
import numpy as np
from dash_iconify import DashIconify
import dash_mantine_components as dmc
import dash_ag_grid as dag

from dash import dcc, html
import plotly.express as px
import plotly.graph_objects as go

from typing import Optional

        
from dash import Output, Input, State, no_update, dcc
import pandas as pd
import numpy as np
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


from .empty_state import render_segments_empty_state





from data import (
    load_columns_df,
    load_df_from_redis,
    delete_df_from_redis,
    save_df_to_redis,
)
from components import MonthSlider, DATES, COLORS_BY_SHADE, COLORS_BY_COLOR
from dash import dcc, Input, Output, State, no_update, MATCH
from .db_queries import get_items, fletch_dataset, fletch_agents, fletch_stores
from .drawler import CATS_MANAGEMENT
from .ag_modal import AGModal
AG_MODAL = AGModal()

COLS = [
    "date",
    "eom",
    "init_date",
    "parent_cat",
    "parent_cat_id",
    "cat",
    "cat_id",
    "subcat",
    "subcat_id",
    "item_id",
    "fullname",
    "brend",
    "manu",
    "amount",
    "quant",
]


# Настройки бенчмарков
# Настройки бенчмарков
OWN_BRANDS      = {"Cosmo", "Cosmo Red"}   # свои бренды
GOOD_WIDTH      = (80, 400)                # глобальная "норма" широты SKU
TAIL_OK_PCT     = 20                       # хвост 50% SKU ≤ 20% выручки — ок (м/д)
HHI_HIGH        = 35                       # высокая концентрация брендов по HHI
TOP1_OK_IF_OWN  = 45                       # целевая доля топ-бренда, если он свой



def derive_good_width_local(
    df_all: pd.DataFrame,
    sku_col: str,            # full_col
    subcat_val,              # str | None
    cat_val,                 # str | None
    min_low: int = 10,
    widen: int = 5
) -> tuple[int, int]:
    """
    Возвращает (low, high) — «нормальную» широту для текущего уровня,
    рассчитанную по пирам (Q25–Q75) среди соседних subcat внутри cat.
    Если данных мало — возвращает глобальный GOOD_WIDTH.
    """
    try:
        if df_all is None or df_all.empty or sku_col not in df_all.columns:
            return GOOD_WIDTH

        peers = df_all.copy()

        # скоуп по родителю (cat), если задан
        if cat_val is not None and "cat" in peers.columns:
            peers = peers.loc[peers["cat"] == cat_val]

        # исключаем техзначения
        bad_vals = {"Нет категории", "Нет подкатегории", "Нет данных", "Нет группы"}
        if "subcat" in peers.columns:
            peers = peers.loc[~peers["subcat"].isin(bad_vals)]

        if "subcat" not in peers.columns:
            return GOOD_WIDTH

        # широта у пиров: уникальные SKU в каждой subcat
        peer_widths = peers.groupby("subcat")[sku_col].nunique()
        peer_widths = peer_widths.dropna()

        # если пиров слишком мало — fallback
        if peer_widths.empty or subcat_val is None or len(peer_widths) < 4:
            return GOOD_WIDTH

        q25 = int(np.quantile(peer_widths, 0.25))
        q75 = int(np.quantile(peer_widths, 0.75))
        low  = max(min_low, q25 - widen)
        high = max(low + 1, q75 + widen)
        return (low, high)
    except Exception:
        return GOOD_WIDTH  # безопасный fallback






def title_badge(text: str):
    return dmc.Badge(
        text,
        color="blue",
        size="lg",
        radius="sm",
        # variant="gradient",
        # gradient={"from": "blue", "to": "grape", "deg": 45},
        style={"fontWeight": 600, "letterSpacing": "0.2px"}
    )
    


def kpi_card(label: str, value: str, icon: str = "mdi:chart-line"):
    return dmc.Paper(
        withBorder=True,
        radius="md",
        shadow="sm",
        p="md",
        style={
            "borderColor": "rgba(99,102,241,0.35)",  # оттенок grape
        },
        children=dmc.Group(
            align="center",
            gap="sm",
            children=[
                dmc.ThemeIcon(
                    size="lg", radius="md", variant="light", color="blue",
                    children=DashIconify(icon=icon, width=20)
                ),
                dmc.Stack(gap=2, children=[
                    dmc.Text(value, size="xl", fw=700, style={"lineHeight": 1}),
                    dmc.Text(label, size="sm", opacity=0.7),
                ])
            ]
        )
    )




def id_to_months(start, end):
    return DATES[start].strftime("%Y-%m-%d"), DATES[end].strftime("%Y-%m-%d")

def fmt_int(x): 
    try: return f"{int(round(float(x))):,}".replace(",", " ")
    except: return "0"

def fmt_rub(x):
    try:
        v = float(x)
        s = f"{v:,.0f}" if abs(v) >= 100 else f"{v:,.2f}"
        return s.replace(",", " ") + " ₽"
    except: return "0 ₽"

def fmt_pct(x):
    try: return f"{float(x):.2f}%".replace(".", ",")
    except: return "0,00%"


    
NBSP_THIN = "\u202F"  # тонкий неразрывный пробел
def fmt_grouped(v: float, money=False):
    try: v = float(v)
    except: v = 0.0
    s = f"{v:,.0f}" if abs(v) >= 100 else f"{v:,.2f}"
    s = s.replace(",", " ").replace(" ", NBSP_THIN)
    return f"{s}{NBSP_THIN}₽" if money else s

_UNITS = ["", "тыс", "млн", "млрд", "трлн"]

def fmt_compact(v: float, money=False, digits=2):
    try: v = float(v)
    except: v = 0.0
    n = abs(v)
    k = 0
    while n >= 1000 and k < len(_UNITS)-1:
        n /= 1000.0
        v /= 1000.0
        k += 1
    num = f"{v:.{digits}f}".rstrip("0").rstrip(".")
    num = num.replace(".", ",")
    out = f"{num}{NBSP_THIN}{_UNITS[k]}".strip()
    return f"{out}{NBSP_THIN}₽" if money else out







def pareto_block(df: pd.DataFrame):
    x = df.copy()

    # safe amount
    if "amount" not in x.columns:
        x["dt"] = pd.to_numeric(x.get("dt"), errors="coerce").fillna(0.0)
        x["cr"] = pd.to_numeric(x.get("cr"), errors="coerce").fillna(0.0)
        x["amount"] = x["dt"] - x["cr"]

    g = (
        x.groupby("fullname", as_index=False)
         .agg(amount=("amount", "sum"))
         .sort_values("amount", ascending=False)
         .reset_index(drop=True)
    )

    total = float(g["amount"].sum()) if not g.empty else 0.0
    g["share"] = (g["amount"] / total * 100).fillna(0) if total > 0 else 0.0
    g["cum_share"] = g["share"].cumsum()

    top = g.head(30).copy()
    top["rank"] = np.arange(1, len(top) + 1)

    chart_data = top[["fullname", "amount", "cum_share"]].to_dict("records")

    NBSP = "\u202F"
    def rub(v):
        try:
            s = f"{float(v):,.0f}".replace(",", " ")
            s = s.replace(" ", NBSP)
            return f"{s}{NBSP}₽"
        except Exception:
            return f"0{NBSP}₽"

    def pct(v):
        try:
            return f"{float(v):.1f}%".replace(".", ",")
        except Exception:
            return "0,0%"

    table = dmc.Table(
        striped=True,
        highlightOnHover=True,
        withTableBorder=True,
        withColumnBorders=False,
        horizontalSpacing="sm",
        verticalSpacing="xs",
        stickyHeader=True,
        children=[
            dmc.TableThead(
                dmc.TableTr([
                    dmc.TableTh("#", style={"width": 44, "textAlign": "right"}),
                    dmc.TableTh("Номенклатура"),
                    dmc.TableTh("Выручка", style={"textAlign": "right", "whiteSpace": "nowrap"}),
                    dmc.TableTh("Доля", style={"textAlign": "right", "whiteSpace": "nowrap"}),
                    dmc.TableTh("Накопл.", style={"textAlign": "right", "whiteSpace": "nowrap"}),
                ])
            ),
            dmc.TableTbody([
                dmc.TableTr([
                    dmc.TableTd(str(int(r.rank)), style={"textAlign": "right", "opacity": 0.8}),
                    dmc.TableTd(dmc.Text(str(r.fullname), lineClamp=1, style={"maxWidth": 520})),
                    dmc.TableTd(rub(r.amount), style={"textAlign": "right", "whiteSpace": "nowrap"}),
                    dmc.TableTd(pct(r.share), style={"textAlign": "right", "whiteSpace": "nowrap"}),
                    dmc.TableTd(pct(r.cum_share), style={"textAlign": "right", "whiteSpace": "nowrap"}),
                ])
                for r in top.itertuples(index=False)
            ])
        ]
    )


    tabs = dmc.Tabs(
        value="chart",
        children=[
            dmc.TabsList([
                dmc.TabsTab("График", value="chart", leftSection=DashIconify(icon="mdi:chart-bar", width=16)),
                dmc.TabsTab("Список", value="list", leftSection=DashIconify(icon="mdi:format-list-bulleted", width=16)),
            ]),
            dmc.TabsPanel(
                value="chart",
                children=[
                    dmc.Space(h=8),
                    dmc.BarChart(
                        h=320,
                        data=chart_data,
                        dataKey="fullname",
                        series=[{"name": "amount", "label": "Выручка"}],
                        valueFormatter={"function": "formatNumberIntl"},
                        withLegend=False,
                        gridAxis="xy",
                    ),
                    dmc.Space(h=8),
                    dmc.Progress(
                        value=float(top["cum_share"].iloc[-1] if not top.empty else 0),
                        size="lg", striped=True, radius="xs"
                    ),
                ],
            ),
            dmc.TabsPanel(
                value="list",
                children=[
                    dmc.Space(h=8),
                    dmc.ScrollArea(
                        h=360,
                        offsetScrollbars=True,
                        type="auto",
                        children=table,
                    ),
                ],
            ),
        ],
    )

    return dmc.Card(
        withBorder=True, radius=0, p="md",
        children=[
            dmc.Group(
                justify="space-between",
                align="center",
                children=[
                    dmc.Title("Парето по выручке (топ-30)", order=5),
                    dmc.Group(gap="xs", children=[
                        dmc.Button(
                            "Excel",
                            id="pareto-export-xlsx",
                            leftSection=DashIconify(icon="mdi:file-excel", width=18),
                            variant="light",
                            color="green",
                            radius="sm",
                            size="sm",
                        ),
                        dcc.Download(id="pareto-download-xlsx"),
                    ]),
                ],
            ),
            dmc.Space(h=6),
            tabs,  # ✅ теперь переменная существует
        ]
    )



# Блок по магазинам/менеджерам
def stores_block(df_stores: pd.DataFrame):
    dfs = df_stores.copy()

        
    dfs['rank'] = dfs['amount'].rank(method='min', ascending=False)
    dfs['store_sales'] = np.where(dfs['rank'] <= 20, dfs['store_gr_name'], 'Другие магазины')
    dfs = dfs.pivot_table(index='store_sales', values='amount', aggfunc='sum').reset_index().sort_values(by='amount', ascending=False)

    store_data = []
    for i, row in enumerate(dfs.itertuples(index=False)):
        color = COLORS_BY_SHADE[i % len(COLORS_BY_SHADE)]  # чтобы не выйти за пределы
        store_data.append({
            "name": row.store_sales,
            "value": float(row.amount),  # на всякий случай преобразуем в число
            "color": color
        })
    
    stores_list = dmc.List(
        [
            dmc.ListItem(
                f"{name}: {value/1_000_000:,.2f} млн ₽".replace(",", " "),
                icon=dmc.ThemeIcon(
                    DashIconify(icon="tdesign:shop-filled", width=16),
                    size=24,
                    radius="xl",
                    color=color_shop,
                    variant="light",
                ),
            )
            for name, value, color_shop in zip(dfs['store_sales'], dfs['amount'], COLORS_BY_SHADE)
        ]
    )
    
    
    return dmc.Box(
        [
            dmc.Space(h=6),
            dmc.Divider(label="Магазины", my="md"),
            dmc.Title("Распределение продаж по магазинам", order=5),
            dmc.Space(h=6),
            
            dmc.Group(
                [  
                 dmc.Stack(
                     [
                        
                        stores_list
                     ]
                    ),
                
                    
                    dmc.Stack(
                     [
                        
                        dmc.PieChart(
                        h=200,
                        data=store_data,
                        labelsType="percent",
                        withTooltip=True,
                        tooltipDataSource="segment",
                        mx="auto",
                        strokeWidth=2,
                        withLabels=True
                     )
                     ]
                    ),
                ],
                gap='xl',
                grow=True,
            )
        ],
       
    )



def insights_block(
    df: pd.DataFrame,
    tot_revenue: float,
    agent_share: float,
    selected_category: Optional[str] = None,
    yaxis_max: Optional[float] = None,
    selected_month_end: Optional[pd.Timestamp] = None,
):
    d = df.copy()

    # ================== Утилиты стиля / UI ==================
    CARD_RADIUS = "lg"
    CARD_SHADOW = "sm"
    PAD = "md"

    def title_badge(text: str, badge: str | None = None):
        return dmc.Group(
            justify="space-between", align="center",
            children=[
                dmc.Text(text, fw=700, fz="md", tt="none"),
                dmc.Badge(badge, variant="outline", radius="xs") if badge else dmc.Box()
            ]
        )

    def section_card(title: str, children, badge: str | None = None):
        return dmc.Paper(
            withBorder=True, radius=CARD_RADIUS, shadow=CARD_SHADOW, p=PAD,
            children=[title_badge(title, badge), dmc.Divider(my=8, variant="dashed"), children],
            style={"backdropFilter": "blur(2px)"}
        )

    # фикс-высота + вертикальный скролл для «ровности» карточек в одном ряду
    def section_card_fixed(title: str, body, badge: str | None = None, h: int = 280):
        header = dmc.Group(
            justify="space-between", align="center",
            children=[
                dmc.Text(title, fw=700, fz="md"),
                dmc.Badge(badge, variant="outline", radius="xs") if badge else dmc.Box()
            ]
        )

        return dmc.Paper(
            withBorder=True,
            radius=CARD_RADIUS,
            shadow=CARD_SHADOW,
            p=PAD,
            style={
                "height": h,
                "display": "flex",
                "flexDirection": "column",
                "overflow": "hidden"
            },
            children=[
                header,
                dmc.Divider(my=8, variant="dashed"),
                dmc.ScrollArea(
                    type="auto",
                    scrollbarSize=8,
                    offsetScrollbars=True,
                    style={
                        "flex": 1,
                        "overflowX": "hidden",
                        "overflowY": "auto",
                        "paddingRight": 6
                    },
                    children=body,
                ),
            ],
        )


    def kpi_card(label: str, value: str, icon: str = "mdi:chart-line", hint: str = ""):
        return dmc.Paper(
            withBorder=True, radius=CARD_RADIUS, shadow="xs", p="sm",
            style={"height": "100%"},
            children=[
                dmc.Group(
                    gap="xs", align="center",
                    children=[
                        dmc.ThemeIcon(DashIconify(icon=icon, width=18), variant="light", radius="xl"),
                        dmc.Text(label, size="sm", c="dimmed"),
                    ]
                ),
                dmc.Text(value, fw=800, fz="lg", mt=4),
                dmc.Text(hint, size="xs", c="dimmed", mt=2) if hint else dmc.Box(),
            ]
        )

    # ================== Форматтеры ==================
    def clamp(x: float) -> float:
        try:
            x = float(x)
        except Exception:
            x = 0.0
        return max(0.0, min(100.0, x))

    def fmt_pct(x, digits=1):
        try:
            s = f"{float(x):.{digits}f}%"
        except Exception:
            s = f"{0:.{digits}f}%"
        return s.replace(".", ",")

    def fmt_compact(x, money=False, digits=1):
        try:
            x = float(x)
        except Exception:
            x = 0.0
        absx = abs(x)
        suffix = ""
        val = x
        if absx >= 1_000_000_000:
            val = x / 1_000_000_000
            suffix = " млрд"
        elif absx >= 1_000_000:
            val = x / 1_000_000
            suffix = " млн"
        elif absx >= 1_000:
            val = x / 1_000
            suffix = " тыс"
        s = f"{val:.{digits}f}".replace(".", ",") + suffix
        if money:
            s += " ₽"
        return s

    def find_col(cands):
        lc = {c.lower(): c for c in d.columns}
        for name in cands:
            if name.lower() in lc:
                return lc[name.lower()]
        return None

    # красивый потолок оси Y (ступени 1–2–5)
    def nice_ceil(v: float, pad_ratio: float = 0.15) -> float:
        v = float(v or 0) * (1 + pad_ratio)
        if v <= 0:
            return 1.0
        mag = 10 ** int(np.floor(np.log10(v)))
        base = np.ceil(v / mag)
        step = 1 if base <= 1 else (2 if base <= 2 else (5 if base <= 5 else 10))
        return float(step * mag)

    # ================== Числа и колонки ==================
    for c in ["dt", "cr", "quant_dt", "quant_cr"]:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0.0)
        else:
            d[c] = 0.0

    d["amount"] = d["dt"] - d["cr"]
    d["quant"] = d["quant_dt"] - d["quant_cr"]
    d["ret_pct"] = (d["cr"] / d["dt"].replace(0, pd.NA)).fillna(0) * 100

    date_col  = find_col(["date", "eom", "doc_date", "operation_date", "sale_date", "init_date", "дата", "Дата"])
    init_col  = find_col(["init_date"])
    agent_col = find_col(["agent_name"])
    store_col = find_col(["store_gr_name"])
    cat_col   = find_col(["category", "category_name", "Категория", "Группа", "group_name"])
    brand_col = find_col(["brend", "brand", "бренд"])
    manu_col  = find_col(["manu", "manufacturer", "производитель"])
    full_col  = find_col(["fullname", "sku", "наименование"])

    if date_col is not None and date_col in d.columns:
        d[date_col] = pd.to_datetime(d[date_col], errors="coerce")
    if init_col is not None and init_col in d.columns:
        d[init_col] = pd.to_datetime(d[init_col], errors="coerce")

    # фильтр по категории
    if selected_category and (cat_col is not None) and (cat_col in d.columns):
        d = d.loc[d[cat_col].astype(str) == str(selected_category)].copy()


    
    # === Новые номенклатуры (последние 30 дней от EOM) ===
    new_block = dmc.Box()
    new_items_cnt = 0
    new_amount_sum = 0.0

    if (init_col is not None and init_col in d.columns and
        date_col is not None and date_col in d.columns and
        full_col is not None and full_col in d.columns and
        not d.empty):

        # Определяем конец выбранного месяца (EOM)
        if selected_month_end is not None:
            eom = pd.to_datetime(selected_month_end, errors="coerce")
            eom = eom.to_period("M").to_timestamp("M") if pd.notna(eom) else pd.NaT
        else:
            max_in_df = pd.to_datetime(d[date_col].max(), errors="coerce")
            eom = max_in_df.to_period("M").to_timestamp("M") if pd.notna(max_in_df) else pd.NaT

        if pd.notna(eom):
            # Ровно 30 календарных дней включительно: [EOM-29d; EOM]
            window_start = eom - pd.Timedelta(days=29)
            new_mask = d[init_col].between(window_start, eom, inclusive="both")
            d_new = d.loc[new_mask].copy()

            if not d_new.empty:
                new_items_cnt  = int(d_new[full_col].nunique())
                new_amount_sum = float(d_new["amount"].sum())

                g_new = (
                    d_new.groupby(full_col, as_index=False)
                        .agg(amount=("amount", "sum"), quant=("quant", "sum"))
                        .assign(price=lambda x: x["amount"] / x["quant"].replace(0, pd.NA))
                        .sort_values("amount", ascending=False)
                )

                def item_two_lines(i, name, q, a, p):
                    top = dmc.Text(f"{i}. {name}", size="sm", lineClamp=1, style={"minWidth": 0})
                    q_txt = f"{int(round(q))} шт" if pd.notna(q) else "— шт"
                    a_txt = fmt_compact(a, money=True)
                    p_txt = f"~{fmt_compact(p, money=True)}/ед" if pd.notna(p) and p > 0 else "~—/ед"
                    bottom = dmc.Text(
                        f"Продано: {q_txt} · {a_txt} · {p_txt}",
                        size="xs", c="dimmed", style={"fontStyle": "italic"}
                    )
                    return dmc.ListItem(dmc.Stack(gap=2, children=[top, bottom]))

                new_items = [
                    item_two_lines(i, getattr(r, full_col), float(r.quant), float(r.amount),
                                (float(r.price) if pd.notna(r.price) else None))
                    for i, r in enumerate(g_new.itertuples(index=False), start=1)
                ]

                header_row = dmc.Group(
                    justify="space-between", align="center",
                    children=[
                        dmc.Group(gap="xs", align="center", children=[
                            dmc.Badge(f"{new_items_cnt} новых SKU", variant="light", color="teal", radius="sm"),
                            dmc.Badge(f"Выручка: {fmt_compact(new_amount_sum, money=True)}", variant="outline", color="blue", radius="sm"),
                        ]),
                        dmc.Text(f"{window_start:%d.%m.%Y} — {eom:%d.%m.%Y}", size="xs", c="dimmed"),
                    ]
                )

                new_block = section_card(
                    "Новые номенклатуры (за последние 30 дней от EOM)",
                    dmc.Stack(
                        gap=6,
                        children=[
                            header_row,
                            dmc.ScrollArea(
                                type="auto", scrollbarSize=8, h=160,
                                styles={"viewport": {"overflowX": "hidden"}},
                                children=dmc.List(new_items, withPadding=True, size="sm", spacing="xs"),
                            ),
                        ],
                    ),
                    badge="Fresh"
                )
            else:
                new_block = section_card(
                    "Новые номенклатуры (за последние 30 дней от EOM)",
                    dmc.Alert(
                        f"За период {window_start:%d.%m.%Y}—{eom:%d.%m.%Y} новых SKU не появилось.",
                        color="gray", variant="light", radius="sm"
                    )
                )
        else:
            new_block = section_card(
                "Новые номенклатуры (за последние 30 дней от EOM)",
                dmc.Alert("Не удалось определить конец месяца для слайдера.", color="gray", variant="light", radius="sm")
            )
    else:
        new_block = section_card(
            "Новые номенклатуры (за последние 30 дней от EOM)",
            dmc.Alert("Колонка init_date не найдена — невозможно определить новые SKU.", color="gray", variant="light", radius="sm")
        )


    # ================== Сводные числа ==================
    n_items = int(d[full_col].nunique()) if full_col in d.columns else 0
    n_brand = int(d[brand_col].nunique()) if brand_col in d.columns else 0
    n_manu  = int(d[manu_col].nunique())  if manu_col  in d.columns else 0

    total_dt = float(d["dt"].sum())
    total_cr = float(d["cr"].sum())
    amount   = float(d["amount"].sum())
    q_net    = float(d["quant"].sum())
    avg_price_net = (amount / q_net) if q_net > 0 else 0.0

    revenue_pct = clamp((amount / tot_revenue * 100) if tot_revenue else 0.0)
    ret_coef    = clamp((total_cr / total_dt * 100) if total_dt else 0.0)
    agent_share = clamp(agent_share or 0.0)

    # Доля «дизайнера»
    designer_share = None
    if agent_col is not None and "dt" in d.columns:
        s = d[agent_col].astype(str)
        is_des = (s.str.contains("дизайн", case=False, na=False) |
                  s.str.fullmatch(r"\s*дизайнер\s*", case=False, na=False))
        dt_des = float(d.loc[is_des, "dt"].sum())
        designer_share = clamp((dt_des / total_dt * 100) if total_dt > 0 else 0.0)

    

    # ================== Парето 80/20 ==================

    by_fullname_amt = (
        d.groupby(full_col, as_index=False)["amount"].sum()
    ) if full_col in d.columns else pd.DataFrame(columns=[full_col, "amount"])

    pareto_body = dmc.Text("нет данных", size="sm", c="dimmed")

    if amount > 0 and n_items > 0 and not by_fullname_amt.empty:
        g = by_fullname_amt.sort_values("amount", ascending=False).reset_index(drop=True)
        total_amt = max(1e-9, g["amount"].sum())
        g["cum_share"] = g["amount"].cumsum() / total_amt * 100

        def k_for(pct: float) -> int:
            return int((g["cum_share"] >= pct).idxmax()) + 1 if not g.empty else 0

        k50 = min(k_for(50), n_items)
        k80 = min(k_for(80), n_items)
        k90 = min(k_for(90), n_items)

        current_cum_80 = float(g["cum_share"].iloc[k80-1]) if k80 > 0 else 0.0
        share_of_cat_80 = (k80 / n_items * 100) if n_items else 0.0

        head_amt_80 = float(g["amount"].head(k80).sum())
        tail_cnt_80 = int(max(0, n_items - k80))
        tail_amt_80 = float(max(0.0, total_amt - head_amt_80))
        tail_share_80 = max(0.0, 100.0 - current_cum_80)
        avg_tail_80 = (tail_amt_80 / tail_cnt_80) if tail_cnt_80 > 0 else 0.0

        header_txt = dmc.Text(
            f"Топ-{k80} SKU (~{share_of_cat_80:.1f}%) формируют 80% выручки".replace(".", ","),
            size="sm", fw=600
        )

        def mini(label, value):
            return dmc.Paper(
                withBorder=True, radius="md", p="xs",
                children=[
                    dmc.Text(label, size="xs", c="dimmed"),
                    dmc.Text(value, fw=700),
                ]
            )

        mini_kpis = dmc.SimpleGrid(
            cols={"base": 3}, spacing="xs",
            children=[
                mini("50% выручки", f"{k50} SKU"),
                mini("80% выручки", f"{k80} SKU"),
                mini("90% выручки", f"{k90} SKU"),
            ]
        )

        bar = dmc.Stack(
            gap=6,
            children=[
                dmc.Progress(value=clamp(current_cum_80), size="md", radius="sm", striped=True, animated=True),
                dmc.Group(justify="space-between", align="center", children=[
                    dmc.Text(f"Достигнуто: {fmt_pct(current_cum_80)}", size="xs", c="dimmed"),
                    dmc.Badge("Цель: 80%", variant="outline", radius="xs"),
                ]),
            ],
        )

        tail_block = dmc.SimpleGrid(
            cols={"base": 2}, spacing="xs",
            children=[
                dmc.Text(f"Хвост: {tail_cnt_80} SKU · {fmt_pct(tail_share_80,1)} выручки", size="xs", c="dimmed"),
                dmc.Text(f"Средняя выручка на 1 SKU в хвосте: {fmt_compact(avg_tail_80, money=True)}", size="xs", c="dimmed"),
            ],
        )

        # === Аналитический вывод (адаптирован под мебель/декор) ===
        if share_of_cat_80 < 18:
            summary_text = dmc.Text(
                "⚠ Критическая концентрация: 80% выручки дают <18% SKU. Сильная зависимость от ограниченного набора хитов.",
                size="xs", c="red", fw=500
            )
        elif share_of_cat_80 < 25:
            summary_text = dmc.Text(
                "⚠ Высокая концентрация: 18–25% SKU дают 80% выручки. Проверь риски по наличию и поставкам топ-позиций.",
                size="xs", c="orange", fw=500
            )
        # elif share_of_cat_80 <= 45:
        #     summary_text = dmc.Text(
        #         "✅ Оптимальное распределение: 25–45% SKU формируют 80% выручки — баланс фокуса и широты ассортимента.",
        #         size="xs", c="green", fw=500
        #     )
        
        elif share_of_cat_80 <= 45:
            if n_items > GOOD_WIDTH[1]:
                summary_text = dmc.Text(
                    "ℹ Структура продаж здоровая (Парето ок), но ассортимент чрезмерно широк — чистка хвоста повысит эффективность.",
                    size="xs", c="orange", fw=500
                )
            else:
                summary_text = dmc.Text(
                    "✅ Оптимальное распределение: 25–45% SKU формируют 80% выручки — баланс фокуса и широты ассортимента.",
                    size="xs", c="green", fw=500
                )
        elif share_of_cat_80 <= 55:
            summary_text = dmc.Text(
                "ℹ Рассеяние продаж: 45–55% SKU для 80% выручки. Вероятен перегруз хвостом — проверь дубли и витрину.",
                size="xs", c="orange", fw=500
            )
        else:
            summary_text = dmc.Text(
                "⚠ Перерассеяние: >55% SKU нужны, чтобы набрать 80% выручки. Сильный хвост — оптимизируй ассортимент.",
                size="xs", c="red", fw=500
            )

        pareto_body = dmc.Stack(
            gap=8,
            children=[
                header_txt,
                mini_kpis,
                bar,
                dmc.Divider(variant="dashed"),
                tail_block,
                summary_text,
            ],
        )


    # pareto_card = section_card_fixed("Парето 80/20", pareto_body, h=260)





    # ================== Топы брендов/производителей ==================
    def top_money(col, k=3):
        if (col is None) or (col not in d.columns) or d.empty:
            return ["нет данных"]
        g = (d.groupby(col, as_index=False)["amount"].sum()
               .sort_values("amount", ascending=False).head(k))
        if g.empty:
            return ["нет данных"]
        return [f"{r[col]} — {fmt_compact(r['amount'], money=True)}" for _, r in g.iterrows()]

    brands_list = top_money(brand_col, k=5)
    manus_list  = top_money(manu_col, k=5)

    # ================== Возвраты/качество ==================
    def risk_color(p):
        return "red" if p >= 30 else ("yellow" if p >= 20 else "green")

    risky_list = [dmc.ListItem("нет данных")]
    risky_count = 0
    very_high_count = 0
    max_ret = 0.0

    if total_dt > 0:
        min_dt = 0.01 * total_dt
        risky_all = d.loc[
            ((d["quant_dt"] >= 5) | (d["dt"] >= min_dt)) & (d["ret_pct"] >= 20)
        ].copy()

        if not risky_all.empty:
            risky_count = int((risky_all["ret_pct"] >= 20).sum())
            very_high_count = int((risky_all["ret_pct"] >= 30).sum())
            max_ret = float(risky_all["ret_pct"].max())

            risky_top = (risky_all.groupby(full_col, as_index=False)
                        .agg(dt=("dt","sum"), cr=("cr","sum"))
                        .assign(ret_pct=lambda x: (x["cr"] / x["dt"].replace(0, pd.NA)).fillna(0) * 100)
                        .sort_values("ret_pct", ascending=False)
                        .head(5))
            risky_list = [
                dmc.ListItem(
                    dmc.Group(
                        gap="xs",
                        children=[
                            dmc.Badge(f"{fmt_pct(r['ret_pct'], 1)}", color=risk_color(float(r["ret_pct"])), variant="filled", radius="xs"),
                            dmc.Text(f"{r[full_col]}", size="sm")
                        ]
                    )
                ) for _, r in risky_top.iterrows()
            ]

    def quality_state(max_ret_val, count_20, count_30):
        if count_20 == 0:
            return ("green", "Возвраты в норме — критичных позиций нет.")
        if count_30 > 0:
            return ("red", f"Высокие возвраты: {count_30} SKU с ≥30%")
        return ("yellow", f"Есть возвраты выше нормы: {count_20} SKU с ≥20%")

    quality_color, quality_text = quality_state(max_ret, risky_count, very_high_count)
    quality_alert = dmc.Alert(
        dmc.Group(gap="xs", children=[DashIconify(icon="mdi:alert-decagram-outline", width=18), dmc.Text(quality_text)]),
        color=quality_color, variant="light", withCloseButton=True, radius="md"
    )

    # ================== Гистограмма цен (опционально) ==================
    price_hist_card = dmc.Box()
    if "quant_dt" in d.columns and (d["quant_dt"] > 0).any():
        d_price = d.loc[(d["quant_dt"] > 0) & (d["dt"] > 0)].copy()
        d_price["price_gross_unit"] = d_price["dt"] / d_price["quant_dt"]
        d_price = d_price.replace([float("inf"), float("-inf")], pd.NA).dropna(subset=["price_gross_unit"])
        if not d_price.empty:
            q_lo, q_hi = d_price["price_gross_unit"].quantile([0.01, 0.99])
            mask = d_price["price_gross_unit"].between(q_lo, q_hi, inclusive="both")
            clipped = int((~mask).sum())
            d_clip = d_price.loc[mask].copy()

            fig_hist = px.histogram(
                d_clip, x="price_gross_unit", nbins=24, labels={"price_gross_unit": "Цена за единицу (Gross)"},
            )
            fig_hist.update_xaxes(range=[float(q_lo), float(q_hi)])
            fig_hist.update_layout(height=220, margin=dict(l=10, r=10, t=10, b=10), hovermode="x unified")

            caption = dmc.Text(f"Показаны 1–99 перцентили; исключено выбросов: {clipped}", size="xs", c="dimmed")
            price_hist_card = section_card(
                "Распределение цен (Gross, за ед.)",
                dmc.Stack(gap=6, children=[dcc.Graph(figure=fig_hist, config={"displayModeBar": False}), caption]),
                badge="Distribution"
            )

    # ================== KPI ==================
    kpis_grid = dmc.SimpleGrid(
        cols={"base": 1, "sm": 2},
        spacing="lg",
        verticalSpacing="lg",
        style={"alignItems": "stretch", "justifyContent": "center"},
        children=[
            kpi_card("Выручка (нетто)", fmt_compact(amount, money=True), icon="mdi:cash-multiple"),
            kpi_card("Средняя цена", fmt_compact(avg_price_net, money=True), icon="mdi:currency-rub"),
            kpi_card("Сумма возвратов", fmt_compact(total_cr, money=True), icon="mdi:cash-refund"),
            kpi_card("Коэфф. возвратов", fmt_pct(ret_coef, 1), icon="mdi:percent"),
        ],
    )

   
    
    # ================== Кольца (индикаторы) + автоанализ ==================

    # плитка кольца (единообразный вид)
    def ring_tile(label: str, value_pct: float, color: str):
        return dmc.Stack(
            align="center", gap=2,
            children=[
                dmc.RingProgress(
                    sections=[{"value": float(clamp(value_pct)), "color": color}],
                    size=88, thickness=10
                ),
                dmc.Text(label, size="xs", c="dimmed"),
                dmc.Text(fmt_pct(value_pct, 2), fw=600),
            ],
            style={"minWidth": 110}
        )

    # сами кольца в ровной сетке (адаптивно)
    rings_grid = dmc.SimpleGrid(
        # cols={"base": 2, "sm": 3, "lg": 4},
        cols=3,
        spacing="lg",
        style={"alignItems": "center", "justifyItems": "center"},
        children=[
            ring_tile("Доля в выручке", revenue_pct, "blue"),
            ring_tile("Коэф. возвратов", ret_coef, "orange"),
            ring_tile("Доля дизайнеров", agent_share, "cyan"),
            (ring_tile("«Дизайнер»", (designer_share or 0), "grape") if designer_share is not None else dmc.Box()),
        ],
    )

    # ---- автоанализ (короткий вердикт менеджеру) ----
    notes = []

    # 1) вклад сегмента в общую выручку
    if revenue_pct >= 30:
        notes.append("Высокий вклад сегмента в выручку — приоритетен для роста/поддержки.")
    elif revenue_pct >= 10:
        notes.append("Средний вклад в выручку — потенциал для масштабирования при сохранении маржи.")
    else:
        notes.append("Низкий вклад — имеет смысл точечно оптимизировать ассортимент/цену.")

    # 2) возвраты
    if ret_coef >= 30:
        ret_level, ret_color = "критично высокие возвраты", "red"
        notes.append("Возвраты критично высокие — проверь качество, логистику и ожидания клиентов.")
    elif ret_coef >= 20:
        ret_level, ret_color = "высокие возвраты", "orange"
        notes.append("Возвраты выше нормы — проведи разбор топ-SKU по обороту и браку.")
    elif ret_coef >= 10:
        ret_level, ret_color = "умеренные возвраты", "yellow"
        notes.append("Возвраты умеренные — держите на контроле проблемные позиции.")
    else:
        ret_level, ret_color = "в норме", "green"
        notes.append("Возвраты в норме — серьёзных рисков не видно.")

    # 3) доля дизайнеров
    if agent_share >= 25:
        notes.append("Высокая доля дизайнеров — зависимость от канала, проверь концентрацию по агентам.")
    elif agent_share <= 5:
        notes.append("Низкая доля дизайнеров — можно тестировать стимулирование канала дизайнеров.")

    # 4) если есть «дизайнер» как агент
    if designer_share is not None:
        if designer_share >= 20:
            notes.append("Сегмент сильно зависит от SKU «дизайнеров» — оцени устойчивость спроса.")
        elif designer_share <= 5:
            notes.append("Вклад SKU «дизайнеров» невелик — есть пространство для промо/витрины.")

    # соберём вывод в Alert (цвет — по возвратам)
    analysis_alert = dmc.Alert(
        dmc.Stack(
            gap=4,
            children=[dmc.Text("Вывод:", fw=600, size="sm")] + [dmc.Text(f"• {t}", size="xs") for t in notes]
        ),
        color=ret_color, variant="light", radius="md"
    )

    # итоговое содержимое карточки «Индикаторы»
    indicators_body = dmc.Stack(
        gap=12,
        children=[
            rings_grid,
            dmc.Divider(variant="dashed"),
            analysis_alert,
        ],
    )
    
    def section_card_fixed_scroll(title: str, body, h: int = 260, badge: str | None = None):
        header = dmc.Group(
            justify="space-between",
            align="center",
            children=[
                dmc.Text(title, fw=700),
                (dmc.Badge(badge, variant="outline", radius="xs") if badge else dmc.Box())
            ],
        )

        return dmc.Card(
            withBorder=True, radius="md", shadow="sm",
  
            style={
                "height": h,
                "display": "flex",
                "flexDirection": "column",
                "overflow": "hidden",  # чтобы внутренности не выпирали за радиус
            },
            children=[
                dmc.CardSection(dmc.Box(p="md", children=header)),
                dmc.Divider(variant="dashed"),
                # ТЕЛО КАРТОЧКИ: занимает остаток высоты
                dmc.CardSection(
                    inheritPadding=True,
                    style={
                        "flex": "1 1 auto",
                        "minHeight": 0,      # ключ к рабочему 100% внутри
                        "display": "flex",
                    },
                    children=dmc.ScrollArea(
                        type="scroll",
                        offsetScrollbars=True,
                        # h="100%",            # 100% от секции выше
                        children=dmc.Box(p="md", children=body),
                    ),
                ),
            ],
        )




    


    # ================== SKU-аналитика ==================
    best_qty_txt = ["нет данных"]
    by_fullname_qty = (
        d.groupby(full_col, as_index=False).agg(q=("quant", "sum"), a=("amount", "sum"))
    ) if full_col in d.columns else pd.DataFrame(columns=[full_col, "q", "a"])
    if not by_fullname_qty.empty:
        bq = by_fullname_qty.sort_values("q", ascending=False).head(5)
        best_qty_txt = []
        for _, r in bq.iterrows():
            price = (r["a"] / r["q"]) if r["q"] > 0 else 0.0
            best_qty_txt.append(f"{r[full_col]} — {int(round(r['q']))} шт, ~{fmt_compact(price, money=True)}/ед")

    best_amt_list = ["нет данных"]
    if not by_fullname_amt.empty:
        top_amt = (
            by_fullname_amt
            .copy()
            .assign(amount=lambda x: pd.to_numeric(x["amount"], errors="coerce").fillna(0.0))
            .query("amount > 0")
            .sort_values("amount", ascending=False)
            .head(5)
        )
        if not top_amt.empty:
            best_amt_list = [f"{r[full_col]} — {fmt_compact(r['amount'], money=True)}" for _, r in top_amt.iterrows()]

    sku_row = dmc.SimpleGrid(
        cols={"base": 1, "sm": 2},
        spacing="md",
        children=[
            section_card("Лучшие SKU по выручке", dmc.List([dmc.ListItem(x) for x in best_amt_list], withPadding=True, size="sm")),
            section_card("Лучшие SKU по количеству", dmc.List([dmc.ListItem(x) for x in best_qty_txt], withPadding=True, size="sm")),
        ],
    )

    # ================== Доп. графики ==================
    def has_children(node) -> bool:
        return getattr(node, "children", None) not in (None, [], ())

    extras_charts = dmc.Box(children=price_hist_card) if has_children(price_hist_card) else dmc.Box()
    
    
    
    
    # ================== Ассортимент  ==================



    def safe_pct(num, den): 
        return (float(num) / float(den) * 100) if den else 0.0

    assort_text = dmc.Group(
        gap="sm",
        children=[
            dmc.Badge(f"{n_items} SKU", color="gray", variant="outline", radius="sm"),
            dmc.Badge(f"{n_brand} брендов", color="gray", variant="outline", radius="sm"),
            dmc.Badge(f"{n_manu} производителей", color="gray", variant="outline", radius="sm"),
        ],
    )


    # 1) Агрегации по брендам и SKU
    dom_brand, top1_share, top3_share, hhi, avg_sku_per_brand = "—", 0.0, 0.0, 0.0, 0.0
    if (brand_col is not None) and (brand_col in d.columns) and not d.empty:
        g_brand = (d.groupby(brand_col, as_index=False)
                    .agg(amount=("amount","sum"),
                        sku_cnt=(full_col, "nunique") if full_col in d.columns else ("amount","size"))
                    .sort_values("amount", ascending=False))
        total_amt_b = float(g_brand["amount"].sum()) or 1e-9
        if not g_brand.empty:
            dom_brand   = str(g_brand.iloc[0][brand_col])
            top1_share  = safe_pct(g_brand.iloc[0]["amount"], total_amt_b)
            top3_share  = safe_pct(g_brand["amount"].head(3).sum(), total_amt_b)
            shares      = (g_brand["amount"] / total_amt_b).to_numpy()
            hhi         = float((shares**2).sum()) * 100  # 0..100
            avg_sku_per_brand = float(g_brand["sku_cnt"].sum() / max(1, len(g_brand)))

    tail50_share = 0.0
    if (full_col is not None) and (full_col in d.columns) and not d.empty:
        g_sku = (d.groupby(full_col, as_index=False)["amount"].sum()
                .sort_values("amount", ascending=False).reset_index(drop=True))
        total_amt_all = float(g_sku["amount"].sum()) or 1e-9
        k50           = int(np.ceil(len(g_sku) * 0.5))
        tail50_amt    = float(g_sku["amount"].iloc[k50:].sum()) if k50 < len(g_sku) else 0.0
        tail50_share  = safe_pct(tail50_amt, total_amt_all)
    
    
    # --- после вычисления tail50_share / pareto_* и ПЕРЕД BLOATED ---

    # дефолт на случай отсутствия пиров
    LOW_W, HIGH_W = GOOD_WIDTH

    # источник пиров: если есть общий d_all — используем его, иначе текущий d
    df_all_source = globals().get("d_all", d)

    curr_subcat = d["subcat"].dropna().unique()[0] if "subcat" in d.columns and d["subcat"].nunique()==1 else None
    curr_cat    = d["cat"].dropna().unique()[0]    if "cat"    in d.columns and d["cat"].nunique()==1    else None

    # попытка получить локальные пороги (если не получится — останутся дефолтные)
    LOW_W, HIGH_W = derive_good_width_local(
        df_all=df_all_source,
        sku_col=full_col,
        subcat_val=curr_subcat,
        cat_val=curr_cat,
        min_low=10,
        widen=5
    )


        # --- Парето-метрика для согласованности с карточкой Парето + признаки "раздут" ---
    pareto_share_80 = None          # % SKU, дающих 80% выручки
    zero_cnt = 0
    zero_share = 0.0
    head_avg_80 = 0.0
    tail_avg_80 = 0.0

    if (full_col is not None) and (full_col in d.columns) and not d.empty:
        g_p = (d.groupby(full_col, as_index=False)["amount"].sum()
                .sort_values("amount", ascending=False).reset_index(drop=True))
        total_p = float(g_p["amount"].sum()) or 1e-9
        g_p["cum_share"] = g_p["amount"].cumsum() / total_p * 100
        zero_cnt = int((g_p["amount"] <= 1e-9).sum())
        zero_share = (zero_cnt / n_items * 100) if n_items else 0.0

        if not g_p.empty:
            k80_idx = int((g_p["cum_share"] >= 80).idxmax())
            k80_loc = k80_idx + 1
            pareto_share_80 = (k80_loc / n_items * 100) if n_items else None

            head_amt_80 = float(g_p["amount"].head(k80_loc).sum()) if k80_loc > 0 else 0.0
            tail_cnt_80 = max(0, n_items - k80_loc)
            tail_amt_80 = max(0.0, total_p - head_amt_80)
            head_avg_80 = (head_amt_80 / k80_loc) if k80_loc > 0 else 0.0
            tail_avg_80 = (tail_amt_80 / tail_cnt_80) if tail_cnt_80 > 0 else 0.0


    # Флаг "раздут"
    BLOATED = (
        n_items > HIGH_W and (
            (pareto_share_80 is not None and pareto_share_80 > 50)
            or (tail50_share >= 25)
            or (zero_share >= 25)
            or (head_avg_80 > 0 and tail_avg_80 < 0.3 * head_avg_80)
        )
    )


    
    # источник пиров: если есть общий d_all — используем его, иначе текущий d
    df_all_source = globals().get("d_all", d)

    curr_subcat = d["subcat"].dropna().unique()[0] if "subcat" in d.columns and d["subcat"].nunique()==1 else None
    curr_cat    = d["cat"].dropna().unique()[0]    if "cat"    in d.columns and d["cat"].nunique()==1    else None

    LOW_W, HIGH_W = derive_good_width_local(
        df_all=df_all_source,
        sku_col=full_col,
        subcat_val=curr_subcat,
        cat_val=curr_cat,
        min_low=10,
        widen=5
    )



    # 2) Статус (светофор) — короткая интегральная оценка
    score = 70

    # --- широта ассортимента ---
    if n_items < LOW_W:
        score -= 15
    elif n_items > HIGH_W:
        if BLOATED:
            score -= 20
        else:
            score -= 5
    else:
        score += 5


    # --- концентрация брендов ---
    if dom_brand in OWN_BRANDS:
        if top1_share >= TOP1_OK_IF_OWN:
            score += 10
        elif top1_share <= 35:
            score -= 5
    else:
        if top1_share >= 45 or hhi >= HHI_HIGH:
            score -= 15

    # --- хвост 50% SKU ---
    if tail50_share <= TAIL_OK_PCT:
        score += 5   # было +10
    elif tail50_share >= 25:
        score -= 10

    # --- доп. штраф за балласт ---
    if zero_share >= 25:
        score -= 10
    if pareto_share_80 is not None and pareto_share_80 > 55:
        score -= 5

    # --- финальный статус ---
    status, color = (
        ("Хорошо", "green") if score >= 75 else
        ("Внимание", "yellow") if score >= 55 else
        ("Риск", "red")
    )

    # Жёсткое правило: раздутый ассортимент не может быть "Хорошо"
    if BLOATED and status == "Хорошо":
        status, color = "Внимание", "yellow"

    status_banner = dmc.Alert(
        dmc.Group(
            gap="xs",
            children=[
                DashIconify(
                    icon=(
                        "mdi:check-decagram" if color == "green"
                        else "mdi:alert" if color == "yellow"
                        else "mdi:alert-octagon"
                    ), width=18
                ),
                dmc.Text(f"Статус: {status}", fw=700),
                dmc.Badge(f"Score {int(score)}", variant="outline", radius="xs"),
            ],
        ),
        color=color, variant="light", radius="md"
    )


    # 3) Вертикальные метрики с вердиктом
    def verdict_badge(text: str, color: str):
        return dmc.Badge(text, color=color, variant="light", radius="sm")


    def label_badge(text: str):
        return dmc.Badge(
            text.upper(),
            variant="light", color="blue", radius="sm", size="sm",
            styles={"root": {"letterSpacing": "0.3px"}}
        )

    def metric_row(label: str, value: str, verdict_text: str, color: str, hint: str = ""):
        return dmc.Group(
            justify="space-between", align="center", style={"width": "100%"},
            children=[
                # ЛЕВАЯ ЧАСТЬ: бейдж-лейбл + значение + хинт
                dmc.Group(
                    gap="sm", align="center",
                    children=[
                        label_badge(label),                               # ← бейдж заголовка
                        dmc.Stack(gap=0, children=[
                            dmc.Text(value, fw=800, fz="lg"),             # значение
                            (dmc.Text(hint, size="xs", c="dimmed")        # хинт
                            if hint else dmc.Box()),
                        ]),
                    ],
                ),
                # ПРАВАЯ ЧАСТЬ: вердикт
                dmc.Badge(verdict_text.upper(), color=color, variant="light", radius="sm"),
            ],
        )


    rows = []
    own = dom_brand in OWN_BRANDS

    # Топ-1 бренд
    if own:
        if top1_share >= TOP1_OK_IF_OWN: v_text, v_color, hint = "норма", "green",  "для собственного бренда целевой фокус"
        elif top1_share >= 40:           v_text, v_color, hint = "внимание","yellow","доля ниже целевой — усилить промо/дистрибуцию"
        else:                            v_text, v_color, hint = "риск",   "red",   "слишком низкая доля собственного бренда"
    else:
        if   top1_share <= 35:           v_text, v_color, hint = "норма",  "green", "зависимости от одного бренда нет"
        elif top1_share <= 45:           v_text, v_color, hint = "внимание","yellow","усиливается зависимость от лидера"
        else:                            v_text, v_color, hint = "риск",   "red",   "высокая зависимость от одного бренда"
    rows.append(metric_row("Топ-1 бренд", f"{dom_brand} · {fmt_pct(top1_share,1)}", v_text, v_color, hint))

    # Топ-3 бренда
    if own and top1_share >= TOP1_OK_IF_OWN:
        v_text, v_color, hint = "норма", "green", "ожидаемо высокая доля топ-3 при собственном бренде"
    else:
        if   top3_share < 55:  v_text, v_color, hint = "внимание","yellow","слишком рассеянная выручка по брендам"
        elif top3_share <= 85: v_text, v_color, hint = "норма",  "green", "здоровая доля топ-3"
        else:                  v_text, v_color, hint = "внимание","yellow","концентрация высоковата — следите за рисками"
    rows.append(metric_row("Топ-3 бренда", fmt_pct(top3_share,1), v_text, v_color, hint))

    # HHI брендов
    if   hhi > 50:
        if own: v_text, v_color, hint = "норма","green","монобрендовая модель — ок для собственного бренда"
        else:   v_text, v_color, hint = "риск", "red",  "высокая концентрация брендов — зависимость"
    elif hhi < 15: v_text, v_color, hint = "норма","green","низкая концентрация, структура диверсифицирована"
    else:          v_text, v_color, hint = "внимание","yellow","умеренная концентрация — держите под контролем"
    rows.append(metric_row("HHI брендов", f"{hhi:.1f}".replace(".", ","), v_text, v_color, hint))

    # SKU на бренд (средняя глубина)
    if   avg_sku_per_brand < 6:   v_text, v_color, hint = "внимание","yellow","малая глубина линейки на бренд"
    elif avg_sku_per_brand <= 50: v_text, v_color, hint = "норма",  "green", "достаточная глубина"
    else:                         v_text, v_color, hint = "внимание","yellow","перегруз по SKU на бренд — проверьте дубли"
    rows.append(metric_row("SKU/бренд (ср.)", f"{avg_sku_per_brand:.1f}".replace(".", ","), v_text, v_color, hint))

    # «Хвост» 50% SKU
    if   tail50_share <= TAIL_OK_PCT: v_text, v_color, hint = "норма","green","тонкий хвост — фокус на топ-SKU"
    elif tail50_share <= 25:          v_text, v_color, hint = "внимание","yellow","умеренный хвост — мониторьте маржу и остатки"
    else:                             v_text, v_color, hint = "риск", "red",  "толстый хвост — чистка/кластеризация промо"
    rows.append(metric_row("Хвост 50% SKU", fmt_pct(tail50_share,1), v_text, v_color, hint))

    metrics_list = dmc.Stack(gap=10, children=rows)
   
    metrics_scroll = dmc.ScrollArea(
    h=120,  # фиксированная высота
    type="scroll",  # показывать вертикальный скролл
    offsetScrollbars=True,
    children=metrics_list,
    style={"paddingRight": "8px"}  # небольшой отступ, чтобы не обрезался скролл
)

    # 4) Короткие выводы (макс. 4)
    bullets = []

    # 1) Широта/раздутие — самое приоритетное
    if n_items < LOW_W:
        bullets.append("Ассортимент узкий — расширить 10–20 позиций в растущих категориях.")
    elif n_items > HIGH_W:
        if BLOATED:
            bullets.append("Ассортимент раздут — убрать низкооборачиваемые и нулепродажные SKU, сократить дубли.")
        else:
            bullets.append("Широта выше среднего для данного уровня, но распределение здоровое — следим за дублями.")
    else:
        bullets.append("Широта ассортимента в норме для данного уровня.")


    # 2) Концентрация по брендам
    if own:
        bullets.append("Собственный бренд — основной драйвер продаж, концентрация допустима."
                    if top1_share >= TOP1_OK_IF_OWN
                    else "Доля собственного бренда ниже целевой — усилить промо и дистрибуцию.")
    else:
        bullets.append("Высокая зависимость от одного бренда — добавить альтернативы."
                    if (top1_share >= 45 or hhi >= HHI_HIGH)
                    else "Концентрация по брендам сбалансирована.")

    # 3) Хвост ассортимента
    if tail50_share <= TAIL_OK_PCT:
        bullets.append("Хвост тонкий — фокусируйтесь на топ-SKU и их наличиях.")
    elif tail50_share >= 25:
        bullets.append("Хвост длинный — проверьте эффективность слабых позиций.")
    else:
        bullets.append("Хвост сбалансирован, без явных рисков.")

    # 4) HHI пояснение
    if hhi > 50:
        bullets.append(f"HHI {hhi:.1f} — {'монобрендовый ассортимент (норма для собственного бренда)' if own else 'высокая концентрация, риск зависимости'}.")
    elif hhi < 15:
        bullets.append(f"HHI {hhi:.1f} — низкая концентрация, структура диверсифицирована.")
    else:
        bullets.append(f"HHI {hhi:.1f} — умеренная концентрация, баланс фокуса и устойчивости.")

    # максимум 4 пункта, по приоритету, как добавляли
    bullets = bullets[:4]

    assort_analysis = dmc.Stack(
        gap=6,
        children=[
            status_banner,
            dmc.List([dmc.ListItem(dmc.Text(x, size="sm")) for x in bullets], withPadding=True, spacing="xs"),
        ],
    )

    # 5) Сборка карточки — без изменений
    assort_card = section_card(
        "Ассортимент",
        dmc.Stack(gap="sm", children=[assort_text, metrics_scroll, dmc.Divider(variant="dashed"), assort_analysis]),
        badge="Structure"
    )





    # ================== Итоговые блоки для «ровности» ==================
    # ---- 1) общий helper для карточек фикс-высоты
    EQUAL_CARD_H = 200  

    def section_card_fixed(title: str, body, h: int = EQUAL_CARD_H, badge: str | None = None):
        return dmc.Paper(
            withBorder=True, radius="lg", shadow="sm", p="md",
            # каркас одинаковой высоты
            style={"height": h, "display": "flex", "flexDirection": "column"},
            children=[
                # шапка
                dmc.Group(
                    justify="space-between", align="center",
                    children=[
                        dmc.Text(title, fw=700, fz="md"),
                        (dmc.Badge(badge, variant="outline", radius="xs") if badge else dmc.Box()),
                    ],
                ),
                dmc.Divider(my=8, variant="dashed"),
                # контентная область заполняет остаток
                dmc.Box(
                    style={"flex": "1 1 auto", "minHeight": 0},  # minHeight=0 позволяет ScrollArea занять место
                    children=body,
                ),
            ],
        )
        

    
    # 1) Левая/правая колонка без индикаторов и без Парето
    left_col = dmc.Stack(
        gap="md",
        children=[
            section_card("KPI", kpis_grid, badge="Overview",),
            # section_card("Ассортимент", dmc.Text(f"{n_items} номенклатур · {n_brand} брендов · {n_manu} производителей", size="sm", opacity=0.8)),
            # assort_card,
        ]
    )

    right_col = dmc.Stack(
        gap="md",
        children=[
            new_block,
        ]
    )

    # 2) Ряд с фикс-высотой: Индикаторы + Парето (ровно, одинаковая высота)
    indicators_card = section_card_fixed_scroll("Индикаторы", indicators_body, h=300)

    pareto_card = section_card_fixed("Парето 80/20", pareto_body, h=300)

    analytics_row = dmc.SimpleGrid(
        cols={"base": 1, "md": 2},
        spacing="lg",
        style={"alignItems": "stretch", "gridAutoRows": "1fr"},
        children=[indicators_card, pareto_card],
    )

    # 3) Топы на всю ширину
    top_brands_manus_fullwidth = dmc.SimpleGrid(
        cols={"base": 1, "sm": 2},
        spacing="md",
        children=[
            section_card("Топ-бренды по выручке", dmc.List([dmc.ListItem(x) for x in brands_list], withPadding=True, size="sm")),
            section_card("Топ-производители по выручке", dmc.List([dmc.ListItem(x) for x in manus_list], withPadding=True, size="sm")),
        ],
    )

    # ================== Хедер ==================
    header = dmc.Group(
        justify="space-between", align="center",
        children=[
            dmc.Group(gap="xs", align="center", children=[
                DashIconify(icon="mdi:flash", width=22),
                dmc.Title("Быстрые выводы", order=4),
            ]),
            dmc.Badge("Аналитика ассортимента", variant="outline", radius="xs"),
        ]
    )

    # ================== Return ==================
    return dmc.Stack(
        gap="md",
        children=[
            header,
            dmc.SimpleGrid(cols={"base": 1, "md": 2}, spacing="lg", children=[left_col, right_col]),
            assort_card,
            analytics_row,  # ← Индикаторы + Парето ровной высоты
            section_card("Рисковые SKU", dmc.Stack(gap=6, children=[quality_alert, dmc.List(risky_list, withPadding=True)])),
            sku_row,
            top_brands_manus_fullwidth,
            extras_charts,
        ],
    )







class SegmentMainWindow:
    def __init__(self):

        self.title = dmc.Title("Сегментный анализ", order=1, c="blue")
        self.memo = dmc.Text(
            "Данный раздел предоставляет аналитику по номенклатурам продукции",
            size="xs",
        )

        self.search_input_id      = "segments_search_fullname"


        self.mslider_id = "segment_analisys_monthslider"
        self.tree_conteiner_id = "segment_analisys_tree_container_very_unique_id"
        self.details_conteiner_id = "segment_analisys_details_container_very_unique_id"
        self.tree_id = "segments_tree_id"
        self.df_store_id = "df_segment_store"
        self.last_update_lb_id = "last_update_segments_lb"
        self.group_box_id = "segment_group_box_id_unique"
        self.assing_cat = "segments_assign_cat_action_button"
        self.assing_manu = "segments_assign_manu_action_button"
        self.assing_brend = "segments_assign_brend_action_button"
        self.ag_grid_id = {'type': 'segments_ag_grid', 'index': '1'}

        self.mslider = MonthSlider(id=self.mslider_id)
        self.tree = dmc.Tree(
            id=self.tree_id,
            data=[],
            expandedIcon=DashIconify(icon="line-md:chevron-right-circle", width=20),
            collapsedIcon=DashIconify(icon="line-md:arrow-up-circle", width=20),
            checkboxes=True,
        )

        group_box_options = [
            {"value": "parent_cat", "label": "Группа"},
            {"value": "brend", "label": "Бренд"},
            {"value": "manu", "label": "Производитель"},
        ]

        self.group_box = dmc.SegmentedControl(
            id=self.group_box_id,
            data=group_box_options,
            value="parent_cat",
        )

        self.df_store = dcc.Store(id=self.df_store_id, storage_type="session")

        self.last_update_lb = dcc.Loading(
            dmc.Badge(
                size="md",
                variant="light",
                radius="xs",
                color="red",
                id=self.last_update_lb_id,
            )
        )
        
        # Делаем иконки для действий с базой данных
        self.actions = dmc.ActionIconGroup(
            [
                dmc.ActionIcon(
                    variant="default",
                    size="lg",
                    children=DashIconify(icon="mdi:file-tree-outline", width=20),
                    id = self.assing_cat,   
                    disabled=True,                 
                ),
                dmc.ActionIcon(
                    variant="default",
                    size="lg",
                    children=DashIconify(icon="mdi:teddy-bear", width=20),
                    id = self.assing_brend,
                    disabled=True, 
                ),
                dmc.ActionIcon(
                    variant="default",
                    size="lg",
                    children=DashIconify(icon="mdi:manufacturing", width=20),
                    id = self.assing_manu,
                    disabled=True, 
                ),
            ],
            orientation="horizontal",
            
            
        )

    def update_ag(self, df, rrgrid_className):
        df = df.copy()
        # безопасные расчёты
        df["dt"]       = pd.to_numeric(df.get("dt"), errors="coerce").fillna(0)
        df["cr"]       = pd.to_numeric(df.get("cr"), errors="coerce").fillna(0)
        df["quant_dt"] = pd.to_numeric(df.get("quant_dt"), errors="coerce").fillna(0)
        df["quant_cr"] = pd.to_numeric(df.get("quant_cr"), errors="coerce").fillna(0)

        df["amount"]    = df["dt"] - df["cr"]
        df["quant"]     = df["quant_dt"] - df["quant_cr"]
        df["ret_ratio"] = (df["cr"] / df["dt"].replace(0, pd.NA)).fillna(0) * 100  # ← ПО ДЕНЬГАМ!

        cols = [
            # ГРУППА «Номенклатура»
            {
                "headerName": "Номенклатура",
                "groupId": "product",
                "marryChildren": True,
                "headerClass": "ag-center-header",
                "children": [
                    {
                        "headerName": "Номенклатура",
                        "field": "fullname",
                        "minWidth": 220,
                        "type": "leftAligned",
                        "cellClass": "ag-firstcol-bg",
                        "headerClass": "ag-center-header",
                         "pinned": "left",
                    },
                    {
                        "headerName": "Бренд",
                        "field": "brend",
                        "minWidth": 160,
                        "type": "leftAligned",
                        "columnGroupShow": "open",
                        "headerClass": "ag-center-header",
                        # опционально:
                        # "valueGetter": {"function": "(p)=> (p.data?.brend && String(p.data.brend).trim()) || 'Бренд не указан'"}
                    },
                    {
                        "headerName": "Производитель",
                        "field": "manu",
                        "minWidth": 160,
                        "type": "leftAligned",
                        "columnGroupShow": "open",
                        "headerClass": "ag-center-header",
                        # "valueGetter": {"function": "(p)=> (p.data?.manu && String(p.data.manu).trim()) || 'Производитель не указан'"}
                    },
                    {
                        "headerName": "Артикль",
                        "field": "article",  # <-- подставь твоё фактическое поле (сейчас у тебя пустая строка)
                        "minWidth": 140,
                        "type": "leftAligned",
                        "columnGroupShow": "open",
                        "headerClass": "ag-center-header",
                    },
                ],
            },

            
            {
                "headerName": "Дата инициализации",
                "field": "init_date",
                "valueFormatter": {"function": "RussianDate(params.value)"},
            },
            {
                "headerName": "Последняя продажа",
                "field": "last_sales_date",
                "valueFormatter": {"function": "RussianDate(params.value)"},
            },
            {
                "headerName": "Выручка",
                "field": "amount",
                "valueFormatter": {"function": "RUB(params.value)"},
                "cellClass": "ag-firstcol-bg",
            },
            {
                "headerName": "Всего продано",
                "field": "quant",
                "valueFormatter": {"function": "FormatWithUnit(params.value,'ед')"},
            },
            {
                "headerName": "Процент возвратов",
                "field": "ret_ratio",
                "valueFormatter": {"function": "FormatWithUnit(params.value,'%')"},
            },
            {"headerName": "Дизайнеры", "field": "agent"},
            {"headerName": "id", "field": "item_id", "hide": True},
]


        return dmc.Stack(
            [
                dmc.Space(h=4),
                dmc.Title("Выбранные позиции", order=4),
                dmc.Space(h=6),
                dag.AgGrid(
                    id=self.ag_grid_id,
                    rowData=df.to_dict("records"),
                    columnDefs=cols,
                    defaultColDef={"sortable": True, "filter": True, "resizable": True},
                    dashGridOptions={
                    "rowSelection": "single", 
                    "pagination": True, 
                    "paginationPageSize": 20,
                    "suppressRowClickSelection": False,
                    #"enableCellTextSelection": True,
                    "ensureDomOrder": True,
                    #"onRowDoubleClicked": {"function": "function(params) { window.dashAgGridFunctions.onRowDoubleClick(params); }"}
                },
                    
                # getRowId="function(params) { return params.data.fullname + '_' + params.data.init_date; }",
                    style={"height": "600px", "width": "100%"},
                    className=rrgrid_className,
                    dangerously_allow_code=True,
                ),
            ]
        )


    def data(self, start_eom, end_eom):
        df = load_columns_df(COLS, start_eom, end_eom)

        return df


    
    
    def maketree(self, df_id, group, top_n=None, min_share=None):
        """
        top_n: показывать у каждого узла не более N самых крупных дочерних веток; остальное -> 'Прочее'
        min_share: скрывать дочерние ветки с долей < min_share (в %) и складывать их в 'Прочее'
        """
        df = load_df_from_redis(df_id)

        # --- подготовка колонок (фикс с *_raw) ---
        if group != "parent_cat":
            df["parent_cat"]    = df[group]
            df["parent_cat_id"] = df[f"{group}_id"]
            df["parent_cat"]    = df["parent_cat"].fillna("Нет данных")
        else:
            df["parent_cat"] = df["parent_cat"].fillna("Нет группы")

        for col, fillv in [
            ("cat", "Нет категории"), ("subcat", "Нет подкатегории")
        ]:
            df[col] = df[col].fillna(fillv)

        df["parent_cat_id"] = df["parent_cat_id"].fillna(10_000_000)
        df["cat_id_raw"]    = df["cat_id"].fillna(10_000_000)
        df["subcat_id_raw"] = df["subcat_id"].fillna(10_000_000)
        df["item_id"]       = df["item_id"].fillna(10_000_001)

        # склейки для стабильного value
        df["cat_id"]    = df["parent_cat_id"].astype(str) + df["cat_id_raw"].astype(str)
        df["subcat_id"] = df["cat_id"] + df["subcat_id_raw"].astype(str)

        # --- метрики ---
        df["amount"] = df.dt - df.cr
        df["quant"]  = df.quant_dt - df.quant_cr

        df = df.pivot_table(
            index=[
                "parent_cat_id","parent_cat",
                "cat_id","cat","cat_id_raw",
                "subcat_id","subcat","subcat_id_raw",
                "fullname","item_id",
            ],
            values=["amount","quant"],
            aggfunc={"amount":"sum","quant":"sum"}
        ).reset_index()

        df["fullname"] = df["fullname"].apply(lambda x: x if len(x) <= 50 else x[:50]+"…")

        # --- helpers ---
        def fmt_money(x):
            ax = abs(x)
            if ax >= 1_000_000_000: return f"{x/1_000_000_000:.1f}млрд"
            if ax >= 1_000_000:     return f"{x/1_000_000:.1f}м"
            if ax >= 1_000:         return f"{x/1_000:.0f}к"
            return f"{x:,.0f}".replace(","," ")

        tree = []

        def find_or_create(lst, value, label):
            for n in lst:
                if n["value"] == str(value): return n
            n = {"value": str(value), "label": str(label), "children": [],
                "_count":0, "_amount":0.0}
            lst.append(n)
            return n

        # --- построение ---
        for _, r in df.iterrows():
            pid, pname = r["parent_cat_id"], r["parent_cat"]
            cid_val, cname = r["cat_id"], r["cat"]
            sid_val, sname = r["subcat_id"], r["subcat"]
            cid_raw, sid_raw = r["cat_id_raw"], r["subcat_id_raw"]
            item_id, fname = r["item_id"], r["fullname"]
            amt = float(r["amount"])

            parent = find_or_create(tree, pid, pname)
            cat    = parent if cid_raw == 10_000_000 else find_or_create(parent["children"], cid_val, cname)
            subcat = cat    if sid_raw == 10_000_000 else find_or_create(cat["children"],    sid_val, sname)

            subcat["children"].append({"value": str(item_id), "label": str(fname)})

            # аккумулируем только по листьям (item_id уникален -> счётчик = уникальные SKU)
            for n in (parent, ) + (() if cat is parent else (cat,)) + (() if subcat in (parent,cat) else (subcat,)):
                n["_count"]  += 1
                n["_amount"] += amt

        # --- сортировка и агрегация "Прочее" ---
        def split_top_rest(children, parent_amount):
            """возвращает (top_children, rest_amount, rest_count)"""
            # интересуют только поддеревья (не листья)
            subs = [c for c in children if isinstance(c, dict) and "children" in c]
            subs.sort(key=lambda n: n.get("_amount",0.0), reverse=True)

            if top_n is None and min_share is None:
                return subs, 0.0, 0

            keep = []
            rest_amt = 0.0
            rest_cnt = 0
            for i, n in enumerate(subs):
                share = (n.get("_amount",0.0) / parent_amount * 100) if parent_amount else 0.0
                cond_top  = (top_n is not None and i < top_n)
                cond_share = (min_share is not None and share >= min_share)
                if (top_n is None and cond_share) or (min_share is None and cond_top) or (cond_top and cond_share):
                    keep.append(n)
                else:
                    rest_amt += n.get("_amount",0.0)
                    rest_cnt += n.get("_count",0)
            return keep, rest_amt, rest_cnt

        def apply_top_rest(node):
            if not node.get("children"): return
            # разделим поддеревья и листья
            subs  = [c for c in node["children"] if isinstance(c, dict) and "children" in c]
            leaves = [c for c in node["children"] if not ("children" in c and isinstance(c["children"], list))]

            top_subs, rest_amt, rest_cnt = split_top_rest(subs, node.get("_amount",0.0))

            # рекурсивно вниз
            for s in top_subs:
                apply_top_rest(s)

            # собираем обратно
            node["children"] = top_subs + leaves

            # добавим "Прочее", если что-то отрезали
            if (rest_amt > 0) or (rest_cnt > 0):
                node["children"].append({
                    "value": f"{node['value']}-others",
                    "label": f"Прочее ({rest_cnt} • ₽{fmt_money(rest_amt)})",
                    "children": []  # лист-агрегатор
                })

        for n in tree:
            apply_top_rest(n)

        # --- финальные подписи (коротко: ₽сумма • N • % ) ---
        total_amount = sum(n.get("_amount",0.0) for n in tree) or None

        def finalize(lst, parent_amount):
            for node in lst:
                amt = float(node.get("_amount", 0.0))
                cnt = int(node.get("_count", 0))

                parts = []
                if cnt:
                    parts.append(str(cnt))                      # 1) количество
                parts.append(f"₽{fmt_money(amt)}")             # 2) выручка
                if parent_amount:                               # 3) доля, если есть родитель
                    share = int(round(amt / parent_amount * 100))
                    parts.append(f"{share}%")

                node["label"] = f"{node['label']} ({' • '.join(parts)})"

                # рекурсия по поддеревьям
                if node.get("children"):
                    subs = [c for c in node["children"] if isinstance(c, dict) and "children" in c]
                    finalize(subs, amt if amt else None)

                # убрать служебные поля
                for k in ("_count", "_amount"):
                    node.pop(k, None)



        finalize(tree, total_amount)
        return tree
    

    
    
    
    
    
    
    def layout(self):
        # --- ЛЕВАЯ КОЛОНКА  ---
        sidebar = dmc.Card(
            withBorder=True, radius="sm", shadow="sm", p="md",
            style={"position": "sticky", "top": 72, "alignSelf": "flex-start"},
            children=[
                dmc.Stack(
                    gap="sm",
                    children=[
                        dmc.Divider(label="Группировка", labelPosition="left"),
                        self.group_box,
                        
                        dmc.Divider(label="Поиск по номенклатуре", labelPosition="left"),
                            dmc.TextInput(
                                id=self.search_input_id,
                                placeholder="Начните вводить название...",
                                leftSection=DashIconify(icon="mdi:magnify", width=18),
                                debounce=350,
                            ),



                        

                        dmc.Divider(label="Выбор позиций", labelPosition="left"),
                        dcc.Loading(children=dmc.ScrollArea(
                            type="scroll",
                            style={"height": 420},
                            children=self.tree,
                        ), type="cube",        ),

                        dmc.Divider(label="Действия", labelPosition="left"),
                        dmc.Flex(self.actions, justify="flex-start"),
                       
                    ],
                )
            ],
        )

        # --- ПРАВАЯ КОЛОНКА (таблица) ---
        right_panel = dmc.Card(
            withBorder=True, radius="sm", shadow="sm", p="md",
            children=[
                dmc.Group(justify="space-between", children=[dmc.Title("Детали / выборка", order=4)]),
                dmc.Space(h=8),
                dcc.Loading(
                    id="segment-details-loading",
                    # type="default",
                    type="cube", 
                    children=dmc.Container(
                        id=self.details_conteiner_id,
                        fluid=True,
            
                         children=render_segments_empty_state(),

                    ),
                ),
            ],
        )

        return dmc.Container(
            fluid=True,
            children=[
                # Заголовок + подзаголовок
                dmc.Group(justify="space-between", align="center",
                        children=[self.title, dmc.Badge("Сегментный анализ", variant="outline", color="blue")]),
                dmc.Text("Данный раздел предоставляет аналитику по номенклатурам продукции", size="xs", c="dimmed"),
                dmc.Space(h=10),

                # --- Ряд "Период" (бейдж справа) ---
                dmc.Group(justify="space-between", align="center",
                        children=[dmc.Text("Период", size="sm", c="dimmed"), self.last_update_lb]),

                # --- СЛАЙДЕР НА ВСЮ ШИРИНУ ---
                dmc.Paper(
                    withBorder=True, radius="sm", p="md",
                    children=self.mslider,  # ← теперь не в колонке, занимает 100%
                ),
                dmc.Space(h=12),
                
                
                # --- ДВЕ КОЛОНКИ ---
                dmc.Grid(
                    gutter="lg", align="stretch",
                    children=[
                        # dmc.GridCol(sidebar, span={"base": 12, "md": 6, "lg": 5, "xl": 4}),
                        # dmc.GridCol(right_panel, span={"base": 12, "md": 6, "lg": 7, "xl": 8}),
                        dmc.GridCol(sidebar, span=4),
                        dmc.GridCol(right_panel, span=8),
                        
                    ],
                ),

                # служебные блоки
                dcc.Store(id="dummy_imputs_for_segment_slider"),
                dcc.Store(id="dummy_imputs_for_segment_render"),
                self.df_store,
                CATS_MANAGEMENT.make_drawler(),
                AG_MODAL.layout()
            ],
        )



    def register_callbacks(self, app):
        
        ag_grid = self.ag_grid_id['type']
        ag_modal = AG_MODAL.modal_id['type']
        model_container = AG_MODAL.modal_conteiner_id['type']
        
        @app.callback(
            Output(self.df_store_id, "data"),
            Output(self.last_update_lb_id, "children"),
            Input(self.mslider_id, "value"),
            Input("dummy_imputs_for_segment_render", "data"),
            State(self.df_store_id, "data"),
            prevent_initial_call=False,
        )
        def update_df(slider_value, dummy, store_data):
            start, end = id_to_months(slider_value[0], slider_value[1])
            start: pd.Timestamp = pd.to_datetime(start) + pd.offsets.MonthBegin(-1)
            end: pd.Timestamp = pd.to_datetime(end) + pd.offsets.MonthEnd(0)

            start = start.strftime("%Y-%m-%d")
            end = end.strftime("%Y-%m-%d")

            if store_data and "df_id" in store_data:
                if store_data["start"] == start and store_data["end"] == end:
                    df = load_df_from_redis(store_data["df_id"])
                    if df is not None:  # ключ ещё живой в Redis
                        min_date = pd.to_datetime(start)
                        max_date = pd.to_datetime(end)
                        notification = f"{min_date.strftime('%b %y')} - {max_date.strftime('%b %y')}"
                        return no_update, notification

                delete_df_from_redis(store_data["df_id"])

            df = fletch_dataset(start, end)
            tot_revenue = df.dt.sum() - df.cr.sum()

            df_id = save_df_to_redis(df, expire_seconds=1200)

            store_dict = {
                "df_id": df_id,
                "start": start,
                "end": end,
                "tot_revenue": tot_revenue,
                "slider_val": slider_value,
            }

            nnoms = df.fullname.nunique()

            min_date = pd.to_datetime(start)
            max_date = pd.to_datetime(end)

            notificattion = f"{min_date.strftime('%b %y')} - {max_date.strftime('%b %y')} ВСЕГО: {nnoms:.0f} НОМЕНКЛАТУР"

            return store_dict, notificattion

        @app.callback(
            Output(self.tree_id, "data"),
            Input(self.df_store_id, "data"),
            Input(self.group_box_id, "value"),
            Input(self.search_input_id, "value"),
        )
        def update_tabs(store_data, group_val, q):
            id_df = store_data["df_id"]
            df = load_df_from_redis(id_df)
            if q:
                mask = df["fullname"].astype(str).str.contains(str(q), case=False, na=False)
                df_f = df.loc[mask].copy()
                tmp_id = save_df_to_redis(df_f, expire_seconds=600)
                out = self.maketree(tmp_id, group_val)
                delete_df_from_redis(tmp_id)
                return out
            return self.maketree(id_df, group_val)

        @app.callback(
            # Output(self.kpi_container_id, "children"),
            Output(self.details_conteiner_id, "children"),
            Output(self.assing_cat, 'disabled'),
            Output(self.assing_brend, 'disabled'),
            Output(self.assing_manu, 'disabled'),
            Input(self.tree_id, "checked"),
            # Input(self.kpi_compact_switch_id, "checked"),
            State("theme_switch", "checked"),
            State(self.df_store_id, "data"),
            prevent_initial_call=True,
        )
        def get_data(checked, theme, store_data):
            if not checked:
                return render_segments_empty_state(), True, True, True
                

            rrgrid_className = "ag-theme-alpine-dark" if theme else "ag-theme-alpine"
            md = get_items(checked, store_data['start'], store_data['end'])

            agent_summary = fletch_agents(checked, store_data['start'], store_data['end'])
            store_summary = fletch_stores(checked, store_data['start'], store_data['end'])
            dfa = agent_summary.copy()
            dfa['agent_sales'] = np.where(dfa['agent_name'] != 'Без дизайнера', "Через дизайнера", dfa['agent_name'])
            dfa = dfa.pivot_table(index='agent_sales', values='amount', aggfunc='sum').reset_index()
            agent_share = dfa[dfa['agent_sales'] == 'Через дизайнера']['amount'].sum() / dfa['amount'].sum() * 100 if dfa['amount'].sum() > 0 else 0.0

            details = dmc.Stack(
                [
                    insights_block(
                        md,
                        store_data['tot_revenue'],
                        agent_share,
                        selected_month_end=pd.to_datetime(store_data['end'])  # <— ключевое
                    ),
                    pareto_block(md),    
                    stores_block(store_summary),
                    self.update_ag(md, rrgrid_className),
                ],
                gap="md",
            )

            return details, False, False, False

        #Вызываем управление категорями
        @app.callback(
            Output(CATS_MANAGEMENT.drawer_id,'opened'),
            Output(CATS_MANAGEMENT.drawer_conteiner_id,'children'),
            Input(self.assing_cat,'n_clicks'),
            State(self.tree_id,'checked'),
            prevent_initial_call=True,            
        )
        def update_cats(nclicks,ids):
            return True, CATS_MANAGEMENT.update_drawer(ids)
                
        @app.callback(
            Output({'type':ag_modal,'index':MATCH},'opened'),
            Output({'type':model_container,'index':MATCH},'children'),
            Input({'type':ag_grid,'index':MATCH},'selectedRows'),
            State(self.df_store_id, "data"),
            prevent_initial_call=True,            
        )
        def open_modal(double_click_data, store_data):
            start = store_data['start']
            end = store_data['end']
            
            if double_click_data:
                
                d = double_click_data[0]
                return True, AG_MODAL.update_modal(d, start, end)
            
            return no_update, no_update
        
        
        

        @app.callback(
            Output("pareto-download-xlsx", "data"),
            Input("pareto-export-xlsx", "n_clicks"),
            State(self.tree_id, "checked"),
            State(self.df_store_id, "data"),
            prevent_initial_call=True,
        )
        def export_pareto_xlsx(n, checked, store_data):
            if not n or not checked or not store_data:
                return no_update

            md = get_items(checked, store_data["start"], store_data["end"])
            if md is None or len(md) == 0:
                return no_update

            # ---------- 1) Подготовка данных ----------
            x = md.copy()

            # дата (для "последней продажи")
            date_col = None
            if "last_sales_date" in x.columns:
                date_col = "last_sales_date"
            elif "date" in x.columns:
                date_col = "date"
            elif "eom" in x.columns:
                date_col = "eom"

            if date_col:
                x[date_col] = pd.to_datetime(x[date_col], errors="coerce")

            # amount (выручка нетто)
            if "amount" not in x.columns:
                x["dt"] = pd.to_numeric(x.get("dt"), errors="coerce").fillna(0.0)
                x["cr"] = pd.to_numeric(x.get("cr"), errors="coerce").fillna(0.0)
                x["amount"] = x["dt"] - x["cr"]
            else:
                x["amount"] = pd.to_numeric(x["amount"], errors="coerce").fillna(0.0)

            # qty (кол-во нетто)
            if "quant" not in x.columns:
                x["quant_dt"] = pd.to_numeric(x.get("quant_dt"), errors="coerce").fillna(0.0)
                x["quant_cr"] = pd.to_numeric(x.get("quant_cr"), errors="coerce").fillna(0.0)
                x["quant"] = x["quant_dt"] - x["quant_cr"]
            else:
                x["quant"] = pd.to_numeric(x["quant"], errors="coerce").fillna(0.0)

            # --------- КЛЮЧ ДЛЯ ГРУППИРОВКИ (ВАЖНО!) ----------
            # чтобы не склеивать разные товары с одинаковым fullname
            has_item_id = "item_id" in x.columns
            key_cols = ["item_id", "fullname"] if has_item_id else ["fullname"]

            # 1) базовая агрегация: выручка, кол-во, средняя цена, последняя дата продажи
            agg = {
                "amount": ("amount", "sum"),
                "qty": ("quant", "sum"),
            }
            if date_col:
                agg["last_sale"] = (date_col, "max")

            g = (
                x.groupby(key_cols, as_index=False)
                .agg(**agg)
                .sort_values("amount", ascending=False)
                .reset_index(drop=True)
            )

            # средняя цена за период (нетто)
            g["avg_price"] = g["amount"] / g["qty"].replace(0, np.nan)

            # доли (0..1)
            total = float(g["amount"].sum()) if not g.empty else 0.0
            g["share"] = (g["amount"] / total) if total > 0 else 0.0
            g["cum_share"] = g["share"].cumsum()

            g.insert(0, "rank", np.arange(1, len(g) + 1))
            top30 = g.head(30).copy()

            # порядок колонок для Excel (БЕЗ "Последней цены продажи")
            cols = ["rank"]
            if has_item_id:
                cols += ["item_id"]
            cols += ["fullname", "amount", "qty", "avg_price"]
            if date_col:
                cols.append("last_sale")
            cols += ["share", "cum_share"]

            rename_map = {
                "rank": "Ранг",
                "fullname": "Номенклатура",
                "amount": "Выручка, ₽",
                "qty": "Кол-во, шт",
                "avg_price": "Средняя цена, ₽",
                "last_sale": "Последняя продажа",
                "share": "Доля выручки, %",
                "cum_share": "Накопленная доля, %",
            }
            if has_item_id:
                rename_map["item_id"] = "ID товара"

            out = top30[cols].rename(columns=rename_map)

            # ---------- 2) Excel через openpyxl ----------
            wb = Workbook()
            ws = wb.active
            ws.title = "Парето (Топ-30)"

            # Запишем датафрейм
            for r in dataframe_to_rows(out, index=False, header=True):
                ws.append(r)

            # Стили
            header_fill = PatternFill("solid", fgColor="1F4E79")  # темно-синий
            header_font = Font(color="FFFFFF", bold=True)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(style="thin", color="D9D9D9")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Шапка
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border

            # Фиксация шапки + фильтр
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # Карточный вид: границы + выравнивание
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    if cell.column == 1:  # Ранг
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        if isinstance(cell.value, (int, float)) or hasattr(cell.value, "year"):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

            # ---------- 3) Форматы чисел по заголовкам ----------
            headers = [c.value for c in ws[1]]
            col_idx = {name: i + 1 for i, name in enumerate(headers)}

            def col_letter(name: str) -> str:
                return get_column_letter(col_idx[name])

            # Деньги
            for name in ("Выручка, ₽", "Средняя цена, ₽"):
                if name in col_idx:
                    L = col_letter(name)
                    for cell in ws[L][1:]:
                        cell.number_format = '#,##0'

            # Кол-во
            if "Кол-во, шт" in col_idx:
                L = col_letter("Кол-во, шт")
                for cell in ws[L][1:]:
                    cell.number_format = '#,##0'

            # Проценты (2 знака, т.к. share 0..1)
            for name in ("Доля, %", "Накопленная доля, %"):
                if name in col_idx:
                    L = col_letter(name)
                    for cell in ws[L][1:]:
                        cell.number_format = '0.00%'

            # Дата
            if "Последняя продажа" in col_idx:
                L = col_letter("Последняя продажа")
                for cell in ws[L][1:]:
                    cell.number_format = 'dd.mm.yyyy'

            # ---------- 4) Зебра ----------
            zebra_fill = PatternFill("solid", fgColor="F6F8FA")
            for r in range(2, ws.max_row + 1):
                if r % 2 == 0:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = zebra_fill

            # ---------- 5) Градиент по выручке ----------
            if "Выручка, ₽" in col_idx:
                L = col_letter("Выручка, ₽")
                ws.conditional_formatting.add(
                    f"{L}2:{L}{ws.max_row}",
                    ColorScaleRule(
                        start_type="min", start_color="E8F5E9",
                        mid_type="percentile", mid_value=50, mid_color="C8E6C9",
                        end_type="max", end_color="81C784",
                    )
                )

            # ---------- 6) Автоширина колонок ----------
            for col in range(1, ws.max_column + 1):
                max_len = 0
                letter = get_column_letter(col)
                for cell in ws[letter]:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            ws.row_dimensions[1].height = 22

            # ---------- 7) Отдаём файл ----------
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            filename = f"Парето_топ30_{store_data['start']}_{store_data['end']}.xlsx"
            return dcc.send_bytes(bio.getvalue(), filename)

        
        
        CATS_MANAGEMENT.register_callbacks(app)
        AG_MODAL.register_callbacks(app)
        
        