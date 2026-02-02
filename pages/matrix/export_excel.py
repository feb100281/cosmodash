# matrix/export_excel.py
from __future__ import annotations

from io import BytesIO
from typing import List, Any, Optional
import json
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# Styles (простые, деловые)
# =========================
TITLE_COLOR = "123A52"
HEADER_COLOR = "DCEAF6"
ZEBRA_COLOR = "F7F9FC"
LINK_FILL = "1F5A7A"
LINK_TEXT = "FFFFFF"

# Акценты для важных колонок
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF4CC")        # мягкий желтый
HIGHLIGHT_HEADER_FILL = PatternFill("solid", fgColor="FFE8A3") # header важн. колонок

thin = Side(style="thin", color="D2D7DD")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

fill_header = PatternFill("solid", fgColor=HEADER_COLOR)
fill_zebra = PatternFill("solid", fgColor=ZEBRA_COLOR)
fill_link = PatternFill("solid", fgColor=LINK_FILL)

font_title = Font(name="Helvetica", size=14, bold=True, color=TITLE_COLOR)
font_subtitle = Font(name="Helvetica", size=11, color="5B6770")
font_header = Font(name="Helvetica", size=11, bold=True, color="1F2D3D")
font_body = Font(name="Helvetica", size=11, color="1F2D3D")
font_link = Font(name="Helvetica", size=11, bold=True, color=LINK_TEXT)

font_toc_child = Font(name="Helvetica", size=11, color="5B6770")

align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)


# =========================
# DB helper
# =========================
def _fetch_barcodes_for_items(engine, item_ids: List[int], start: str, end: str) -> pd.DataFrame:
    """
    Возвращает разрез по штрихкодам сразу по всем item_id.
    Группировка: item_id + barcode.
    """
    if not item_ids:
        return pd.DataFrame(columns=["item_id", "barcode", "amount", "quant", "share"])

    placeholders = ", ".join([f"%(id_{i})s" for i in range(len(item_ids))])
    params = {f"id_{i}": int(v) for i, v in enumerate(item_ids)}
    params.update({"start": start, "end": end})

    q = f"""
        select
            t.item_id as item_id,
            coalesce(b.barcode,'нет штрихкода') as barcode,
            sum(t.dt-t.cr) as amount,
            sum(t.quant_dt - t.quant_cr) as quant
        from sales_salesdata as t
        left join corporate_barcode as b on b.id = t.barcode_id
        where t.item_id in ({placeholders})
          and LAST_DAY(t.date) between %(start)s and %(end)s
        group by t.item_id, b.barcode
        order by t.item_id, b.barcode
    """

    df = pd.read_sql(q, engine, params=params)

    if df.empty:
        df["share"] = []
        return df

    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
    totals = df.groupby("item_id")["amount"].transform("sum")
    df["share"] = df["amount"] / totals.replace({0: 1})

    return df


# =========================
# Utils
# =========================
def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    name = re.sub(r"[:\\/?*\[\]]", " ", str(name)).strip()
    name = re.sub(r"\s+", " ", name)
    if not name:
        name = "Лист"
    return name[:max_len]


def _try_parse_list(val: Any) -> Optional[List[Any]]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, list):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        try:
            obj = json.loads(s)
            if isinstance(obj, list):
                return obj
        except Exception:
            return None
    return None


def _format_date_key(d: Any) -> str:
    try:
        dt = pd.to_datetime(d)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(d)


# =========================
# Excel styling helpers
# =========================
def _style_sheet_basic(
    ws,
    title: str,
    subtitle: str,
    toc_name: str = "Оглавление",
    header_row: int = 4,
    freeze_cell: str = "C5",  # заморозка 2 колонок + верх
):
    """
    Макет:
      row1: Title
      row2: Subtitle
      row3: ссылка "← Оглавление" (без merge)
      row4: header таблицы
      row5+: данные
    """
    ws.sheet_view.showGridLines = False

    ws.insert_rows(1, 3)

    ws["A1"] = title
    ws["A1"].font = font_title
    ws.row_dimensions[1].height = 24

    ws["A2"] = subtitle
    ws["A2"].font = font_subtitle

    # --- Кнопка "← Оглавление" (merge на всю ширину таблицы) ---
    last_col = ws.max_column

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

    cell = ws.cell(3, 1, "← Оглавление")
    cell.hyperlink = f"#{toc_name}!A1"
    cell.font = font_link
    cell.fill = fill_link
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Прокрашиваем и обводим всю merged-строку, чтобы выглядело как кнопка
    for c in range(1, last_col + 1):
        ws.cell(3, c).fill = fill_link
        ws.cell(3, c).border = border_thin

    ws.row_dimensions[3].height = 20


    ws.freeze_panes = freeze_cell

    max_col = ws.max_column
    max_row = ws.max_row

    # Header
    for c in range(1, max_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = border_thin

    # Body + zebra
    for r in range(header_row + 1, max_row + 1):
        zebra = (r - (header_row + 1)) % 2
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.font = font_body
            # серый хинт "раскройте +"
            if cell.value == "раскройте +":
                cell.font = Font(name="Helvetica", size=10, color="9AA3AC", italic=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            cell.border = border_thin
            cell.alignment = align_left
            if zebra:
                cell.fill = fill_zebra

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"


def _apply_number_formats(ws, header_row: int = 4):
    headers = {c: (ws.cell(header_row, c).value or "") for c in range(1, ws.max_column + 1)}

    for c, name in headers.items():
        n = str(name).strip().lower()

        # Штрихкод — всегда текст
        if n in ("штрихкод", "barcode"):
            for r in range(header_row + 1, ws.max_row + 1):
                ws.cell(r, c).number_format = "@"
            continue

        # ₽
        if n in ("выручка", "amount", "сумма", "ср. выручка", "_amount"):
            fmt = '#,##0" ₽"'
            align = align_center
        # qty
        elif n in ("кол-во", "количество", "quant", "qty", "шт", "кол-во (по датам)", "ср. μ (ед)", "ст откл. σ", "cv квар."):
            fmt = "#,##0.00"
            align = align_center
        # %
        elif n in ("доля", "share", "процент", "%", "доля выручки", "доля в ср выручке", "_share", "cum_share"):
            fmt = "0.00%"
            align = align_center
        else:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt
                cell.alignment = align


def _apply_period_month_format(ws, header_row: int = 4):
    """
    Приводит колонки 'Нач. период' и 'Конеч. период' к формату 'Янв. 2026'.
    Работает и если там datetime, и если строка типа 'Июля 2025', 'Авг. 2025', '2025-08-01' и т.п.
    """
    targets = {"нач. период", "конеч. период"}

    month_map = {
        1: "Янв.", 2: "Февр.", 3: "Мар.", 4: "Апр.", 5: "Май",
        6: "Июн.", 7: "Июл.", 8: "Авг.", 9: "Сент.", 10: "Окт.", 11: "Ноя.", 12: "Дек."
    }

    # найдём индексы колонок
    cols = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h and str(h).strip().lower() in targets:
            cols.append(c)

    if not cols:
        return

    for c in cols:
        for r in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v in (None, ""):
                continue

            # пробуем распарсить в дату
            dt = None
            try:
                dt = pd.to_datetime(v, errors="coerce", dayfirst=True)
            except Exception:
                dt = None

            if dt is None or pd.isna(dt):
                continue

            # формируем "Ммм. YYYY"
            m = int(dt.month)
            y = int(dt.year)
            cell.value = f"{month_map.get(m, '')} {y}".strip()
            cell.number_format = "@"
            cell.alignment = align_center



def _set_fixed_widths(
    ws,
    header_row: int = 4,
    numeric_width: int = 14,
    text_width: int = 26,
):
    """
    Единые фиксированные ширины:
      - numeric_width: все числовые колонки + ABC/XYZ
      - text_width: все остальные (текстовые)
    """
    numeric_headers = {
        # классификаторы
        "abc", "xyz",

        # деньги
        "выручка", "amount", "сумма", "ср. выручка", "_amount",

        # количества
        "кол-во", "количество", "quant", "qty", "шт", "кол-во (по датам)",

        # доли / проценты
        "доля", "share", "процент", "%", "доля выручки", "доля в ср выручке", "_share", "cum_share",

        # статистика
        "cv квар.", "ср. μ (ед)", "ст откл. σ", "макс. (ед)", "мин. (ед)",

        # периоды (часто числа)
        "qпер. (мес)", "нулевые периоды (мес)", "периоды с продажами (мес)",

        # запасы
        "страх. запас (ед) (ss)", "rop (ед)",
    }

    for c in range(1, ws.max_column + 1):
        header = ws.cell(header_row, c).value
        if not header:
            continue
        h = str(header).strip().lower()
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = numeric_width if h in numeric_headers else text_width


def _hide_columns_by_headers(ws, header_names: List[str], header_row: int = 4):
    targets = {h.strip().lower() for h in header_names}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        if str(v).strip().lower() in targets:
            ws.column_dimensions[get_column_letter(c)].hidden = True


def _highlight_columns(ws, header_names: List[str], header_row: int = 4):
    targets = {h.strip().lower() for h in header_names}
    cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip().lower() in targets:
            cols.append(c)

    for c in cols:
        hcell = ws.cell(header_row, c)
        hcell.fill = HIGHLIGHT_HEADER_FILL
        hcell.font = Font(name="Helvetica", size=11, bold=True, color="1F2D3D")

        for r in range(header_row + 1, ws.max_row + 1):
            ws.cell(r, c).fill = HIGHLIGHT_FILL


# =========================
# Outline/grouping
# =========================
def _outline_group_children(
    ws,
    marker_header: str,
    indent_header: str,
    header_row: int = 4,
    collapse_by_default: bool = True,
):
    """
    parent row: marker пустой
    child row: marker заполнен
    """
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h:
            col_map[str(h).strip()] = c

    marker_col = col_map.get(marker_header)
    indent_col = col_map.get(indent_header)
    if not marker_col or not indent_col:
        return

    current_parent = None

    for r in range(header_row + 1, ws.max_row + 1):
        marker_val = ws.cell(r, marker_col).value

        if marker_val in (None, ""):
            current_parent = r
            ws.row_dimensions[r].outlineLevel = 0
            ws.row_dimensions[r].collapsed = False
            # parent жирнее
            # parent жирнее (НО не трогаем подсказку "раскройте +")
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value == "раскройте +":
                    continue
                cell.font = Font(name="Helvetica", size=11, bold=True, color="1F2D3D")

        else:
            ws.row_dimensions[r].outlineLevel = 1
            if collapse_by_default:
                ws.row_dimensions[r].hidden = True
                if current_parent:
                    ws.row_dimensions[current_parent].collapsed = True
            ws.cell(r, indent_col).alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)


# =========================
# TOC (простое, без merge)
# =========================
def _build_toc_simple_grouped(
    wb,
    toc_name: str,
    main_sheets: List[str],
    manufacturers_parent: str,
    manufacturer_sheets: List[str],
) -> str:
    if toc_name in wb.sheetnames:
        wb.remove(wb[toc_name])

    ws = wb.create_sheet(toc_name, 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Оглавление"
    ws["A1"].font = Font(name="Helvetica", size=16, bold=True, color=TITLE_COLOR)
    ws.row_dimensions[1].height = 26

    ws["A3"] = "Лист"
    ws["B3"] = "Перейти"
    for cell in (ws["A3"], ws["B3"]):
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = border_thin

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = "A4"

    def add_row(r: int, name: str, child: bool = False) -> int:
        cell_name = ws.cell(r, 1, name)
        cell_link = ws.cell(r, 2, "Открыть →")
        cell_link.hyperlink = f"#{name}!A1"

        if child:
            cell_name.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell_name.font = font_toc_child
            cell_link.font = Font(name="Helvetica", size=11, bold=True, color="6B7A86")
        else:
            cell_name.alignment = align_left
            cell_name.font = font_body
            cell_link.font = Font(name="Helvetica", size=11, bold=True, color="1F5A7A")

        cell_name.border = border_thin
        cell_link.border = border_thin

        if (r - 4) % 2:
            cell_name.fill = fill_zebra
            cell_link.fill = fill_zebra

        return r + 1

    r = 4
    for s in main_sheets:
        if s in wb.sheetnames:
            r = add_row(r, s, child=False)

    if manufacturers_parent in wb.sheetnames:
        r = add_row(r, manufacturers_parent, child=False)
        for s in manufacturer_sheets:
            if s in wb.sheetnames:
                r = add_row(r, s, child=True)

    return toc_name


# =========================
# Builders: manufacturers
# =========================
def _build_manufacturers_sheets(writer: pd.ExcelWriter, df_matrix_export: pd.DataFrame) -> List[str]:
    if "Производитель" not in df_matrix_export.columns:
        return []

    df = df_matrix_export.copy()
    df["Производитель"] = df["Производитель"].fillna("Нет производителя").astype(str).str.strip()

    tmp = df.copy()
    tmp["SKU"] = 1

    sv = tmp.groupby("Производитель", as_index=False).agg(
        SKU=("SKU", "sum"),
        **({"Выручка": ("Выручка", "sum")} if "Выручка" in tmp.columns else {}),
        **({"Кол-во": ("Кол-во", "sum")} if "Кол-во" in tmp.columns else {}),
    )

    if "Выручка" in sv.columns and float(sv["Выручка"].sum()) != 0:
        sv["Доля выручки"] = sv["Выручка"] / sv["Выручка"].sum()
    elif "Выручка" in sv.columns:
        sv["Доля выручки"] = 0.0

    sort_col = "Выручка" if "Выручка" in sv.columns else "SKU"
    sv = sv.sort_values(sort_col, ascending=False)

    sv.to_excel(writer, index=False, sheet_name="Производители")

    # user cols: без служебных, без json, без детальных листов
    user_cols = [c for c in df.columns if c.lower() not in {"item_id", "subcat_id", "cat_id"} and not c.lower().endswith("_json")]
    service_cols = {"item_id", "subcat_id", "cat_id", "ls_quant", "ls_date", "is_quant", "is_date"}

    user_cols = [
        c for c in df.columns
        if c.lower() not in service_cols
        and not c.lower().endswith("_json")
    ]
    # исключаем детальные колонки, которые мы добавляем только для матрицы
    # user_cols = [c for c in user_cols if c not in ("Дата продажи", "Кол-во (по датам)")]

    created = []
    for manu in sv["Производитель"].tolist():
        sub = df[df["Производитель"] == manu][user_cols].copy()
        if sub.empty:
            continue

        sheet = _safe_sheet_name(manu)
        base = sheet
        i = 2
        while sheet in writer.book.sheetnames:
            sheet = _safe_sheet_name(f"{base[:28]} {i}")
            i += 1

        sub.to_excel(writer, index=False, sheet_name=sheet)
        created.append(sheet)

    return created


# =========================
# Матрица: детализация по датам продаж (outline)
# =========================
def _make_matrix_dates_hierarchical(
    df_matrix_export: pd.DataFrame,
    df_raw: pd.DataFrame,
) -> pd.DataFrame:
    """
    parent: строка товара (Штрихкод пустой)
    child: даты продаж + qty + (по одному штрихкоду в строке)
    """
    if "item_id" not in df_raw.columns:
        return df_matrix_export

    # источники для детализации
    date_src = None
    quant_src = None
    for c in df_raw.columns:
        if str(c).lower() in ("ls_date", "is_date"):
            date_src = c
        if str(c).lower() in ("ls_quant", "is_quant"):
            quant_src = c

    if not date_src or not quant_src:
        return df_matrix_export

    raw_map = df_raw[["item_id", date_src, quant_src]].copy()
    raw_map["item_id"] = pd.to_numeric(raw_map["item_id"], errors="coerce")
    raw_map = raw_map.dropna(subset=["item_id"])
    raw_map["item_id"] = raw_map["item_id"].astype(int)
    raw_map = raw_map.drop_duplicates(subset=["item_id"]).set_index("item_id")

    df = df_matrix_export.copy()

    # добавляем колонки для детализации
    if "Дата продажи" not in df.columns:
        df["Дата продажи"] = None
    if "Кол-во (по датам)" not in df.columns:
        df["Кол-во (по датам)"] = None

    cols = list(df.columns)

    out_rows = []
    for _, row in df.iterrows():
        item_id = row.get("item_id")


        # ---------- parent ----------
        parent = row.to_dict()

        hint = "раскройте +"

        # ВАЖНО: "Дата продажи" НЕ заполняем подсказкой,
        # потому что по ней мы определяем parent/child в outline
        if "Дата продажи" in parent:
            parent["Дата продажи"] = None

        if "Штрихкод" in parent:
            parent["Штрихкод"] = hint

        if "Кол-во (по датам)" in parent:
            parent["Кол-во (по датам)"] = hint

        out_rows.append(parent)



        if item_id is None or (isinstance(item_id, float) and pd.isna(item_id)):
            continue

        try:
            item_id_int = int(item_id)
        except Exception:
            continue

        if item_id_int not in raw_map.index:
            continue

        dates = _try_parse_list(raw_map.loc[item_id_int, date_src])
        quants = _try_parse_list(raw_map.loc[item_id_int, quant_src])
        if not dates or not quants:
            continue

        n = min(len(dates), len(quants))
        if n <= 0:
            continue

        # ---------- подготовим пары дата/кол-во ----------
        pairs = []
        for i in range(n):
            d_str = _format_date_key(dates[i])
            try:
                d_dt = pd.to_datetime(d_str)
            except Exception:
                d_dt = pd.NaT
            try:
                q = float(quants[i])
            except Exception:
                q = quants[i]
            pairs.append((d_dt, d_str, q))

        # сортировка по дате
        pairs.sort(key=lambda x: (pd.Timestamp.max if pd.isna(x[0]) else x[0]))

        # ---------- подготовим список штрихкодов (разбиваем строку "..., ...") ----------
        barcodes: List[Optional[str]] = []
        if "Штрихкод" in row and row.get("Штрихкод") not in (None, ""):
            bc_val = row.get("Штрихкод")
            if isinstance(bc_val, str):
                barcodes = [x.strip() for x in bc_val.split(",") if x.strip()]
            else:
                barcodes = [str(bc_val).strip()]
        if not barcodes:
            barcodes = [None]  # чтобы строки по датам не пропадали

        # ---------- children ----------
        for _, d_str, q in pairs:
            for bc in barcodes:
                child = {c: None for c in cols}

                # для визуальной структуры оставляем ключевые поля, чтобы:
                # 1) фильтр по Производителю не убивал детальные строки
                # 2) в детализации было понятно, что за позиция
                for keep_col in ("Номенклатура", "Производитель", "Артикул"):
                    if keep_col in child:
                        child[keep_col] = row.get(keep_col)

                # штрихкод (по одному в строке)
                if "Штрихкод" in child:
                    child["Штрихкод"] = bc

                child["Дата продажи"] = d_str
                child["Кол-во (по датам)"] = q

                # item_id оставляем (он скрыт), чтобы не ломать логику
                child["item_id"] = row.get("item_id")

                out_rows.append(child)



    df_out = pd.DataFrame(out_rows)

    # удобный порядок: "Дата продажи" и "Кол-во (по датам)" ближе к датам
    def _move_after(df_: pd.DataFrame, col_to_move: str, after_col: str) -> pd.DataFrame:
        if col_to_move not in df_.columns or after_col not in df_.columns:
            return df_
        cols_ = list(df_.columns)
        cols_.remove(col_to_move)
        idx = cols_.index(after_col) + 1
        cols_.insert(idx, col_to_move)
        return df_[cols_]

    # сначала ставим Дату продажи
    if "Периоды с продажами (мес)" in df_out.columns:
        df_out = _move_after(df_out, "Дата продажи", "Периоды с продажами (мес)")
    elif "Конеч. период" in df_out.columns:
        df_out = _move_after(df_out, "Дата продажи", "Конеч. период")

    # затем Штрихкод — сразу после Даты продажи
    if "Штрихкод" in df_out.columns and "Дата продажи" in df_out.columns:
        df_out = _move_after(df_out, "Штрихкод", "Дата продажи")

    # и количество — после штрихкода
    if "Кол-во (по датам)" in df_out.columns and "Штрихкод" in df_out.columns:
        df_out = _move_after(df_out, "Кол-во (по датам)", "Штрихкод")


    return df_out


# =========================
# Штрихкоды: иерархия
# =========================
def _make_barcodes_hierarchical(df_barcodes: pd.DataFrame) -> pd.DataFrame:
    if df_barcodes.empty:
        return df_barcodes

    need = ["Номенклатура", "Штрихкод", "Выручка", "Кол-во", "Доля", "item_id"]
    for c in need:
        if c not in df_barcodes.columns:
            df_barcodes[c] = None

    out_rows = []
    for item_id, g in df_barcodes.groupby("item_id", dropna=False):
        g = g.copy()
        name = g["Номенклатура"].iloc[0]

        parent = {
            "Номенклатура": name,
            "Штрихкод": None,
            "Выручка": float(pd.to_numeric(g["Выручка"], errors="coerce").fillna(0).sum()),
            "Кол-во": float(pd.to_numeric(g["Кол-во"], errors="coerce").fillna(0).sum()),
            "Доля": 1.0,
            "item_id": item_id,
        }
        out_rows.append(parent)

        for _, r in g.iterrows():
            out_rows.append({
                "Номенклатура": r.get("Номенклатура"),
                "Штрихкод": r.get("Штрихкод"),
                "Выручка": r.get("Выручка"),
                "Кол-во": r.get("Кол-во"),
                "Доля": r.get("Доля"),
                "item_id": r.get("item_id"),
            })

    return pd.DataFrame(out_rows)


# =========================
# Footnote
# =========================
def _add_footnote(ws, header_row: int = 4):
    row = ws.max_row + 2
    last_col = ws.max_column

    text = (
        "* ABC — вклад в выручку; XYZ — стабильность спроса;\n"
        "  μ — средние продажи в месяц; σ — стандартное отклонение;\n"
        "  CV — коэффициент вариации; SS и ROP — параметры точки заказа"
    )

    # merge на всю ширину таблицы
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)

    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Helvetica", size=11, color="5B6770")
    cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    # высота под 3 строки текста
    ws.row_dimensions[row].height = 48




# =========================
# Main export
# =========================
def build_matrix_excel_bytes(engine, df_matrix: pd.DataFrame, start: str, end: str) -> bytes:
    """
    XLSX:
      - Оглавление (вложенное, без объединений)
      - Матрица (детализация по датам с outline, штрихкод в детальных строках, сортировка по дате)
      - Штрихкоды (outline)
      - Производители + листы производителей (оформлены)
    Лист "Динамика (json)" НЕ создаётся.
    """
    df_raw = df_matrix.copy()
    
    SERVICE_COLS = {
    "item_id", "subcat_id", "cat_id",
    "date_json", "quant_json",
    "ls_quant", "ls_date", "is_quant", "is_date",
}


    if "item_id" not in df_raw.columns:
        raise ValueError("В df_matrix нет колонки 'item_id' — экспорт невозможен.")

    # штрихкоды
    item_ids = df_raw["item_id"].dropna().astype(int).unique().tolist()
    df_barcodes = _fetch_barcodes_for_items(engine, item_ids=item_ids, start=start, end=end)

    if not df_barcodes.empty and "fullname" in df_raw.columns:
        map_name = df_raw[["item_id", "fullname"]].drop_duplicates().set_index("item_id")["fullname"].to_dict()
        df_barcodes["fullname"] = df_barcodes["item_id"].map(map_name)

    # rename columns
    matrix_rename = {
        "fullname": "Номенклатура",
        "article": "Артикул",
        "manu": "Производитель",
        "abc": "ABC",
        "xyz": "XYZ",
        "barcode": "Штрихкод",
        "cat_name": "Категория",
        "sc_name": "Подкатегория",
        "amount": "Выручка",
        "quant": "Кол-во",
        "share": "Доля выручки",
        "mean_amount": "Ср. выручка",
        "share_mean": "Доля в ср выручке",
        "mean_month": "Ср. μ (ед)",
        "std_month": "Ст откл. σ",
        "cv": "CV Квар.",
        "max_month": "Макс. (ед)",
        "min_month": "Мин. (ед)",
        "min_date": "Нач. период",
        "max_date": "Конеч. период",
        "sales_period_months": "Qпер. (мес)",
        "missing_months": "Нулевые периоды (мес)",
        "month_count": "Периоды с продажами (мес)",
        "ss": "Страх. запас (ед) (SS)",
        "rop": "ROP (ед)",
        "item_id": "item_id",
        "subcat_id": "subcat_id",
        "cat_id": "cat_id",
        "date_json": "date_json",
        "quant_json": "quant_json",
        "ls_quant": "ls_quant",
        "ls_date": "ls_date",
        "is_quant": "is_quant",
        "is_date": "is_date",

    }

    df_matrix_export = df_raw.rename(columns={k: v for k, v in matrix_rename.items() if k in df_raw.columns})
    
    # убрать служебные колонки из выгрузки
    df_matrix_export = df_matrix_export.drop(columns=["cum_share", "_amount", "_share"], errors="ignore")


    df_barcodes = df_barcodes.rename(
        columns={
            "fullname": "Номенклатура",
            "barcode": "Штрихкод",
            "amount": "Выручка",
            "quant": "Кол-во",
            "share": "Доля",
            "item_id": "item_id",
        }
    )

    # порядок колонок (Матрица)
    prefer_cols = [
        "ABC", "XYZ",
        "Номенклатура", "Артикул", "Производитель",
        "Категория", "Подкатегория",
        "Штрихкод",
        "Выручка", "Кол-во", "Доля выручки",
        "Ср. выручка", "Доля в ср выручке",
        "Ср. μ (ед)", "Ст откл. σ", "CV Квар.", "Макс. (ед)", "Мин. (ед)",
        "Нач. период", "Конеч. период", "Qпер. (мес)", "Нулевые периоды (мес)", "Периоды с продажами (мес)",
        "Страх. запас (ед) (SS)", "ROP (ед)",
        "cat_id", "subcat_id", "item_id", "date_json", "quant_json", "ls_quant", "ls_date", "is_quant", "is_date",
    ]
    df_matrix_export = df_matrix_export[
        [c for c in prefer_cols if c in df_matrix_export.columns]
        + [c for c in df_matrix_export.columns if c not in prefer_cols]
    ]

    # детализация по датам + по одному штрихкоду в детальной строке
    df_matrix_export = _make_matrix_dates_hierarchical(df_matrix_export, df_raw)
    df_matrix_xlsx = df_matrix_export.drop(columns=[c for c in SERVICE_COLS if c in df_matrix_export.columns], errors="ignore")


    # штрихкоды: иерархия
    df_barcodes = _make_barcodes_hierarchical(df_barcodes)

    # Export
    out = BytesIO()
    manufacturer_sheets: List[str] = []

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_matrix_xlsx.to_excel(writer, index=False, sheet_name="Матрица")
        df_barcodes.to_excel(writer, index=False, sheet_name="Штрихкоды")
        manufacturer_sheets = _build_manufacturers_sheets(writer, df_matrix_xlsx)


    out.seek(0)
    wb = load_workbook(out)

    # TOC
    toc_name = _build_toc_simple_grouped(
        wb,
        toc_name="Оглавление",
        main_sheets=["Матрица", "Штрихкоды", ],
        manufacturers_parent="Производители",
        manufacturer_sheets=manufacturer_sheets,
    )

    # ================
    # Матрица style
    # ================
    ws_m = wb["Матрица"]
    _style_sheet_basic(
        ws_m,
        title="Ассортиментная матрица",
        subtitle=f"Период: {start} — {end}",
        toc_name=toc_name,
        header_row=4,
        freeze_cell="D5",
    )
    _apply_number_formats(ws_m, header_row=4)

    _apply_period_month_format(ws_m, header_row=4)
    _set_fixed_widths(ws_m, header_row=4, numeric_width=12, text_width=26)

    _hide_columns_by_headers(
        ws_m,
        ["item_id", "subcat_id", "cat_id", "date_json", "quant_json", "ls_quant", "ls_date", "is_quant", "is_date"],
        header_row=4,
    )

    _highlight_columns(ws_m, ["Страх. запас (ед) (SS)", "ROP (ед)"], header_row=4)

    _outline_group_children(
        ws_m,
        marker_header="Дата продажи",
        indent_header="Номенклатура",
        header_row=4,
        collapse_by_default=True,
    )

    _add_footnote(ws_m, header_row=4)

    # ================
    # Штрихкоды style
    # ================
    ws_b = wb["Штрихкоды"]
    _style_sheet_basic(
        ws_b,
        title="Детализация по штрихкодам",
        subtitle=f"Период: {start} — {end}",
        toc_name=toc_name,
        header_row=4,
        freeze_cell="C5",
    )
    _apply_number_formats(ws_b, header_row=4)
    _set_fixed_widths(ws_b, header_row=4, numeric_width=12, text_width=26)
    _hide_columns_by_headers(ws_b, ["item_id"], header_row=4)

    _outline_group_children(
        ws_b,
        marker_header="Штрихкод",
        indent_header="Номенклатура",
        header_row=4,
        collapse_by_default=True,
    )

    # =====================
    # Производители sheet
    # =====================
    if "Производители" in wb.sheetnames:
        ws_p = wb["Производители"]
        _style_sheet_basic(
            ws_p,
            title="Производители",
            subtitle=f"Период: {start} — {end}",
            toc_name=toc_name,
            header_row=4,
            freeze_cell="C5",
        )
        _apply_number_formats(ws_p, header_row=4)
        _set_fixed_widths(ws_p, header_row=4, numeric_width=12, text_width=26)

        # добавим колонку "Перейти"
        header_row = 4
        last_col = ws_p.max_column + 1
        ws_p.cell(header_row, last_col, "Перейти").fill = fill_header
        ws_p.cell(header_row, last_col, "Перейти").font = font_header
        ws_p.cell(header_row, last_col, "Перейти").alignment = align_header
        ws_p.cell(header_row, last_col, "Перейти").border = border_thin
        ws_p.column_dimensions[get_column_letter(last_col)].width = 14

        manu_col_idx = None
        for c in range(1, ws_p.max_column + 1):
            if (ws_p.cell(header_row, c).value or "").strip() == "Производитель":
                manu_col_idx = c
                break

        if manu_col_idx:
            for r in range(header_row + 1, ws_p.max_row + 1):
                manu = ws_p.cell(r, manu_col_idx).value
                manu = "Нет производителя" if manu is None else str(manu).strip()
                sheet = _safe_sheet_name(manu)
                if sheet not in wb.sheetnames:
                    continue
                cell = ws_p.cell(r, last_col, "Открыть →")
                cell.hyperlink = f"#{sheet}!A1"
                cell.font = Font(name="Helvetica", size=11, bold=True, color="1F5A7A")
                cell.border = border_thin
                cell.alignment = align_center
                if (r - (header_row + 1)) % 2:
                    cell.fill = fill_zebra

    # =========================
    # Sheets by manufacturer
    # =========================
    for s in manufacturer_sheets:
        if s not in wb.sheetnames:
            continue
        ws = wb[s]
        _style_sheet_basic(
            ws,
            title=f"Производитель: {s}",
            subtitle=f"Период: {start} — {end}",
            toc_name=toc_name,
            header_row=4,
            freeze_cell="C5",
        )
        _outline_group_children(
            ws,
            marker_header="Дата продажи",
            indent_header="Номенклатура",
            header_row=4,
            collapse_by_default=True,
        )
        _apply_number_formats(ws, header_row=4)
        _apply_period_month_format(ws, header_row=4)
        _set_fixed_widths(ws, header_row=4, numeric_width=12, text_width=26)
        _highlight_columns(ws, ["Страх. запас (ед) (SS)", "ROP (ед)"], header_row=4)
        _hide_columns_by_headers(ws, ["item_id", "subcat_id", "cat_id", "date_json", "quant_json"], header_row=4)

    final = BytesIO()
    wb.save(final)
    return final.getvalue()

